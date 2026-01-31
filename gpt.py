import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext, filedialog
import pyodbc

# Check pandas availability for advanced features
try:
    import pandas as pd
    import numpy as np
    PANDAS_AVAILABLE = True
except ImportError:
    print("Warning: pandas/numpy not available. Advanced template mapping features disabled.")
    PANDAS_AVAILABLE = False
    pd = None
    np = None
import json
import os
import logging
from datetime import datetime, timedelta
import threading
import sys
from tkcalendar import DateEntry
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.drawing.image import Image as XLImage
from PIL import Image, ImageTk
import io
import re
import traceback
import subprocess
import platform
import difflib
from typing import Dict, List, Optional, Any, Union, Tuple
from dataclasses import dataclass, field, asdict
from collections import OrderedDict

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('table_exporter.log'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# ============================================================================
# TEMPLATE ANALYZER (AUTO HEADER + POSITION DETECTION)
# ============================================================================

class ExcelTemplateAnalyzer:
    """Analyze an Excel template to detect header rows/columns for each sheet."""

    @staticmethod
    def _looks_like_header(value: Any) -> bool:
        if value is None:
            return False
        s = str(value).strip()
        if not s:
            return False
        # avoid purely numeric headers
        numeric_like = s.replace('.', '').replace(',', '').replace('-', '').isdigit()
        return not numeric_like

    @staticmethod
    def detect_sheet_headers(ws, max_rows: int = 30, max_cols: int = 60) -> Optional[Dict[str, Any]]:
        """
        Detect a header "band" in a sheet by finding a row with 2+ consecutive text-like cells.
        Returns dict: {header_row, start_col, end_col, headers}
        """
        max_row = min(max_rows, ws.max_row or max_rows)
        max_col = min(max_cols, ws.max_column or max_cols)

        best = None
        best_len = 0

        for r in range(1, max_row + 1):
            current_start = None
            current_headers = []
            for c in range(1, max_col + 1):
                v = ws.cell(row=r, column=c).value
                if ExcelTemplateAnalyzer._looks_like_header(v):
                    if current_start is None:
                        current_start = c
                    current_headers.append(str(v).strip())
                else:
                    if current_start is not None and len(current_headers) >= 2:
                        if len(current_headers) > best_len:
                            best_len = len(current_headers)
                            best = {
                                "header_row": r,
                                "start_col": current_start,
                                "end_col": c - 1,
                                "headers": current_headers[:]
                            }
                    current_start = None
                    current_headers = []

            # end-of-row flush
            if current_start is not None and len(current_headers) >= 2:
                if len(current_headers) > best_len:
                    best_len = len(current_headers)
                    best = {
                        "header_row": r,
                        "start_col": current_start,
                        "end_col": current_start + len(current_headers) - 1,
                        "headers": current_headers[:]
                    }

        return best

    @staticmethod
    def analyze_template(template_path: str) -> Dict[str, Dict[str, Any]]:
        """Return analysis per sheet name."""
        wb = load_workbook(template_path, data_only=True, keep_links=False)
        analysis: Dict[str, Dict[str, Any]] = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            info = ExcelTemplateAnalyzer.detect_sheet_headers(ws)
            if info:
                analysis[sheet_name] = info
        return analysis

    @staticmethod
    def _calculate_similarity(text1: str, text2: str) -> float:
        """Calculate similarity between two text strings using various methods."""
        text1 = text1.lower().strip()
        text2 = text2.lower().strip()

        # Exact match gets highest score
        if text1 == text2:
            return 1.0

        # Check for substring matches
        if text1 in text2 or text2 in text1:
            return 0.9

        # Use difflib for fuzzy matching
        return difflib.SequenceMatcher(None, text1, text2).ratio()

    @staticmethod
    def generate_auto_mappings(template_headers: List[str], db_columns: List[str],
                              confidence_threshold: float = 0.6) -> Dict[str, str]:
        """
        Automatically generate mappings between template headers and database columns.

        Args:
            template_headers: List of headers from Excel template
            db_columns: List of available database columns
            confidence_threshold: Minimum similarity score for auto-mapping (0.0-1.0)

        Returns:
            Dict mapping template_header -> db_column
        """
        mappings = {}

        for template_header in template_headers:
            best_match = None
            best_score = 0.0

            for db_col in db_columns:
                similarity = ExcelTemplateAnalyzer._calculate_similarity(template_header, db_col)
                if similarity > best_score and similarity >= confidence_threshold:
                    best_score = similarity
                    best_match = db_col

            if best_match:
                mappings[template_header] = best_match

        return mappings

    @staticmethod
    def analyze_and_map_template(template_path: str, db_columns: List[str],
                                confidence_threshold: float = 0.6) -> Dict[str, Dict[str, Any]]:
        """
        Analyze template and auto-generate mappings for all sheets.

        Returns:
            Dict with structure:
            {
                sheet_name: {
                    'analysis': {...},  # Original analysis data
                    'auto_mappings': {...},  # Auto-generated mappings
                    'confidence_scores': {...}  # Similarity scores for mappings
                }
            }
        """
        analysis = ExcelTemplateAnalyzer.analyze_template(template_path)
        result = {}

        for sheet_name, sheet_info in analysis.items():
            template_headers = sheet_info.get('headers', [])

            # Generate auto mappings
            auto_mappings = ExcelTemplateAnalyzer.generate_auto_mappings(
                template_headers, db_columns, confidence_threshold
            )

            # Calculate confidence scores for each mapping
            confidence_scores = {}
            for template_header, db_col in auto_mappings.items():
                confidence_scores[template_header] = ExcelTemplateAnalyzer._calculate_similarity(
                    template_header, db_col
                )

            result[sheet_name] = {
                'analysis': sheet_info,
                'auto_mappings': auto_mappings,
                'confidence_scores': confidence_scores,
                'unmapped_headers': [h for h in template_headers if h not in auto_mappings]
            }

        return result

# ============================================================================
# DATA CLASSES
# ============================================================================

@dataclass
class TablePosition:
    """Stores position information for a table"""
    table_name: str
    sheet_name: str
    start_row: int = 16  # Minimum 4 to leave space for title
    start_col: str = "A"
    header_positions: Dict[str, str] = None
    merge_ranges: List[str] = None  # optional ranges like "A1:B1"
    template_headers: List[str] = None  # detected headers for mapping preview

    def __post_init__(self):
        if self.header_positions is None:
            self.header_positions = {}
        if self.merge_ranges is None:
            self.merge_ranges = []
        if self.template_headers is None:
            self.template_headers = []
    
    def get_start_col_num(self) -> int:
        """Convert column letter to number"""
        col = self.start_col.upper()
        result = 0
        for char in col:
            result = result * 26 + (ord(char) - ord('A') + 1)
        return result

@dataclass
class CellMapping:
    """Mapping information for a single cell"""
    template_sheet: str
    template_cell: str
    data_source: str  # Could be 'fixed', 'data_table', 'calculated'
    data_key: Optional[str] = None  # Column name or key to fetch data from
    data_table: Optional[str] = None  # Which data table to use
    row_offset: int = 0  # Row offset from template position
    col_offset: int = 0  # Column offset from template position

@dataclass
class TableConfig:
    """Configuration for a data table positioning"""
    sheet_name: str
    start_row: int
    start_col: str
    header_row: Optional[int] = None  # Row where headers are in template
    data_start_row: Optional[int] = None  # Row where data starts in template
    auto_detect: bool = True

    def get_start_col_num(self) -> int:
        """Convert column letter to number"""
        col = self.start_col.upper()
        result = 0
        for char in col:
            result = result * 26 + (ord(char) - ord('A') + 1)
        return result

@dataclass
class TemplateConfig:
    """Complete template configuration"""
    template_name: str
    template_file: str
    cell_mappings: Dict[str, CellMapping] = field(default_factory=dict)
    table_configs: Dict[str, TableConfig] = field(default_factory=dict)
    fixed_values: Dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> Dict:
        return {
            "template_name": self.template_name,
            "template_file": self.template_file,
            "cell_mappings": {k: asdict(v) for k, v in self.cell_mappings.items()},
            "table_configs": {k: asdict(v) for k, v in self.table_configs.items()},
            "fixed_values": self.fixed_values
        }

    @classmethod
    def from_dict(cls, data: Dict) -> 'TemplateConfig':
        config = cls(
            template_name=data["template_name"],
            template_file=data["template_file"]
        )
        config.fixed_values = data.get("fixed_values", {})

        # Load cell mappings
        for key, mapping_data in data.get("cell_mappings", {}).items():
            config.cell_mappings[key] = CellMapping(**mapping_data)

        # Load table configs
        for table_name, table_data in data.get("table_configs", {}).items():
            config.table_configs[table_name] = TableConfig(**table_data)

        return config

# ============================================================================
# EXCEL EXPORTER - FIXED VERSION
# ============================================================================

class ExcelExporter:
    """Export data to Excel with template support"""
    
    def __init__(self):
        self.wb = None
        self.ws = None
        self.current_sheet = None
        self.table_positions = []
        self.column_widths = {}
        
    def export_to_excel(self, tables: Dict[str, pd.DataFrame], output_path: str,
                       template_path: Optional[str] = None,
                       table_positions: List[TablePosition] = None) -> bool:
        """
        Export tables to Excel, optionally using a template
        
        Args:
            tables: Dictionary of table names to DataFrames
            output_path: Path to save the Excel file
            template_path: Optional template file path
            table_positions: Optional list of TablePosition objects for positioning
            
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            logger.info(f"📤 Creating Excel file...")
            
            if template_path and os.path.exists(template_path):
                return self._export_with_template(tables, output_path, template_path, table_positions)
            else:
                return self._export_without_template(tables, output_path)
                
        except Exception as e:
            logger.error(f"❌ Export error: {str(e)}")
            logger.error(traceback.format_exc())
            return False
    
    def _export_with_template(self, tables: Dict[str, pd.DataFrame], output_path: str,
                             template_path: str, table_positions: List[TablePosition]) -> bool:
        """Export using a template"""
        try:
            # Load template
            self.wb = load_workbook(template_path)
            
            # Apply table positions if provided
            if table_positions:
                for table_pos in table_positions:
                    self._place_table_in_template(tables.get(table_pos.table_name), table_pos)
            else:
                # Default behavior: place each table in its own sheet
                for table_name, df in tables.items():
                    if table_name in self.wb.sheetnames:
                        self.ws = self.wb[table_name]
                        self._write_dataframe(df, start_row=4)
                    else:
                        # Create new sheet if it doesn't exist
                        self.ws = self.wb.create_sheet(title=table_name[:31])
                        self._write_dataframe(df, start_row=1, write_headers=True)
            
            # Save the workbook
            self.wb.save(output_path)
            logger.info(f"✅ Excel file created successfully: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"❌ Template export error: {str(e)}")
            return False
    
    def _export_without_template(self, tables: Dict[str, pd.DataFrame], output_path: str) -> bool:
        """Export without template - create new workbook"""
        try:
            self.wb = Workbook()
            
            # Remove default sheet
            if 'Sheet' in self.wb.sheetnames:
                default_sheet = self.wb['Sheet']
                self.wb.remove(default_sheet)
            
            # Create a sheet for each table
            for table_name, df in tables.items():
                # Truncate sheet name to 31 characters (Excel limit)
                sheet_name = table_name[:31]
                self.ws = self.wb.create_sheet(title=sheet_name)
                self.current_sheet = sheet_name
                
                # Write table name as title
                self.ws['A1'] = table_name
                self.ws['A1'].font = Font(bold=True, size=14)
                
                # Write dataframe starting from row 3
                self._write_dataframe(df, start_row=3, write_headers=True)
                
                # Auto-adjust column widths
                self._auto_adjust_columns(df)
            
            # Save the workbook
            self.wb.save(output_path)
            logger.info(f"✅ Excel file created successfully: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"❌ Export error: {str(e)}")
            return False
    
    def _place_table_in_template(self, df: pd.DataFrame, table_pos: TablePosition):
        """Place a DataFrame at specified position in template"""
        if df is None or df.empty:
            logger.warning(f"⚠ No data for table: {table_pos.table_name}")
            return
        
        # Get or create the sheet
        if table_pos.sheet_name in self.wb.sheetnames:
            self.ws = self.wb[table_pos.sheet_name]
        else:
            self.ws = self.wb.create_sheet(title=table_pos.sheet_name[:31])
        
        # Use get_start_col_num() method - FIXED
        start_col_num = table_pos.get_start_col_num()
        
        # Write headers if specified
        if table_pos.header_positions:
            for header, cell_ref in table_pos.header_positions.items():
                if header in df.columns:
                    self.ws[cell_ref] = header
        
        # Write data starting from specified position
        start_row = max(table_pos.start_row, 4)  # Minimum row 4 for safety
        
        # Write the dataframe
        for i, (_, row) in enumerate(df.iterrows()):
            for j, value in enumerate(row):
                cell = self.ws.cell(row=start_row + i, column=start_col_num + j)
                cell.value = value
        
        logger.info(f"📊 Placed table '{table_pos.table_name}' at {table_pos.sheet_name}!{table_pos.start_col}{start_row}")
    
    def _write_dataframe(self, df: pd.DataFrame, start_row: int = 1, write_headers: bool = True):
        """Write DataFrame to current worksheet"""
        if df is None or df.empty:
            return
        
        # Write headers
        if write_headers:
            for col_idx, column_name in enumerate(df.columns, 1):
                cell = self.ws.cell(row=start_row, column=col_idx)
                cell.value = column_name
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            
            start_row += 1
        
        # Write data
        for i, (_, row) in enumerate(df.iterrows()):
            for j, value in enumerate(row, 1):
                cell = self.ws.cell(row=start_row + i, column=j)
                cell.value = value
        
        logger.info(f"📝 Wrote {len(df)} rows to sheet '{self.current_sheet}'")
    
    def _auto_adjust_columns(self, df: pd.DataFrame):
        """Auto-adjust column widths based on content"""
        if df is None or df.empty:
            return
        
        # Calculate max width for each column
        for col_idx, column in enumerate(df.columns, 1):
            max_length = len(str(column))
            
            # Check data in column
            for cell in df[column]:
                if cell is not None:
                    max_length = max(max_length, len(str(cell)))
            
            # Set column width (add some padding)
            adjusted_width = min(max_length + 2, 50)  # Max width 50
            column_letter = get_column_letter(col_idx)
            self.ws.column_dimensions[column_letter].width = adjusted_width

# ============================================================================
# MAIN APPLICATION
# ============================================================================

class DatabaseTableExporter:
    """Main application class"""
    
    def __init__(self):
        self.db_manager = None
        self.exporter = ExcelExporter()
        self.tables_data = {}
        
    def connect_to_database(self, server: str, database: str, username: str, password: str) -> bool:
        """Connect to SQL Server database"""
        try:
            self.db_manager = DatabaseManager()
            success = self.db_manager.connect(server, database, username, password)
            return success
        except Exception as e:
            logger.error(f"❌ Connection error: {e}")
            return False
    
    def fetch_tables(self, table_names: List[str]) -> Dict[str, pd.DataFrame]:
        """Fetch data from specified tables"""
        if not self.db_manager or not self.db_manager.is_connected:
            logger.error("❌ Not connected to database")
            return {}
        
        try:
            all_tables = {}
            total_rows = 0
            
            for table_name in table_names:
                logger.info(f"📋 Fetching data from table: {table_name}")
                
                # Get column names first
                columns = self.db_manager.get_column_names(table_name)
                if not columns:
                    logger.warning(f"⚠ No columns found for table: {table_name}")
                    continue
                
                # Build query
                columns_str = ", ".join([f"[{col}]" for col in columns])
                query = f"SELECT {columns_str} FROM [{table_name}]"
                
                # Execute query
                data = self.db_manager.query_data(query)
                
                if data:
                    # Convert to DataFrame
                    df = pd.DataFrame(data, columns=columns)
                    all_tables[table_name] = df
                    total_rows += len(df)
                    logger.info(f"✅ Fetched {len(df)} rows from {table_name}")
                else:
                    logger.warning(f"⚠ No data found in table: {table_name}")
            
            logger.info(f"✅ Successfully fetched {len(all_tables)} tables ({total_rows} total rows)")
            self.tables_data = all_tables
            return all_tables
            
        except Exception as e:
            logger.error(f"❌ Error fetching tables: {e}")
            return {}
    
    def export_tables(self, output_path: str, template_path: Optional[str] = None,
                     table_positions: Optional[List[TablePosition]] = None) -> bool:
        """Export fetched tables to Excel"""
        if not self.tables_data:
            logger.warning("⚠ No data to export")
            return False
        
        return self.exporter.export_to_excel(
            self.tables_data,
            output_path,
            template_path,
            table_positions
        )
    
    def analyze_template(self, template_path: str) -> Dict[str, Dict[str, Any]]:
        """Analyze an Excel template for headers and structure"""
        if not os.path.exists(template_path):
            logger.error(f"❌ Template file not found: {template_path}")
            return {}
        
        try:
            analyzer = ExcelTemplateAnalyzer()
            analysis = analyzer.analyze_template(template_path)
            
            logger.info(f"✅ Template analysis complete. Found {len(analysis)} sheet(s) with headers:")
            for sheet_name, info in analysis.items():
                logger.info(f"   📊 {sheet_name}: {len(info.get('headers', []))} headers at row {info.get('header_row')}")
            
            return analysis
            
        except Exception as e:
            logger.error(f"❌ Template analysis error: {e}")
            return {}
    
    def generate_auto_mappings(self, template_path: str, confidence_threshold: float = 0.6) -> Dict[str, Dict[str, Any]]:
        """Generate automatic mappings between template and database columns"""
        if not self.db_manager or not self.db_manager.is_connected:
            logger.error("❌ Not connected to database")
            return {}
        
        try:
            # Get all database columns from all tables
            all_columns = []
            table_names = self.db_manager.get_table_names()
            
            for table_name in table_names:
                columns = self.db_manager.get_column_names(table_name)
                all_columns.extend(columns)
            
            # Remove duplicates
            all_columns = list(set(all_columns))
            logger.info(f"📊 Found {len(all_columns)} unique database columns")
            
            # Analyze template and generate mappings
            analyzer = ExcelTemplateAnalyzer()
            result = analyzer.analyze_and_map_template(
                template_path,
                all_columns,
                confidence_threshold
            )
            
            # Log results
            total_mappings = 0
            total_unmapped = 0
            
            for sheet_name, sheet_info in result.items():
                mappings = sheet_info.get('auto_mappings', {})
                unmapped = sheet_info.get('unmapped_headers', [])
                
                total_mappings += len(mappings)
                total_unmapped += len(unmapped)
                
                logger.info(f"   📋 {sheet_name}: {len(mappings)} mapped, {len(unmapped)} unmapped headers")
            
            logger.info(f"✅ Auto-mapping complete: {total_mappings} total mappings, {total_unmapped} unmapped headers")
            
            return result
            
        except Exception as e:
            logger.error(f"❌ Auto-mapping error: {e}")
            return {}

# ============================================================================
# GUI APPLICATION
# ============================================================================

class TableExporterGUI:
    """GUI for Database Table Exporter"""
    
    def __init__(self, root):
        self.root = root
        self.root.title("Database Table Exporter")
        self.root.geometry("900x700")
        
        self.exporter = DatabaseTableExporter()
        self.setup_ui()
        
    def setup_ui(self):
        """Setup the user interface"""
        # Create main notebook
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Database Connection Tab
        self.setup_connection_tab()
        
        # Table Selection Tab
        self.setup_table_tab()
        
        # Template Tab
        self.setup_template_tab()
        
        # Export Tab
        self.setup_export_tab()
        
        # Log Tab
        self.setup_log_tab()
        
        # Status bar
        self.status_var = tk.StringVar(value="Ready")
        status_bar = ttk.Label(self.root, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W)
        status_bar.pack(side=tk.BOTTOM, fill=tk.X)
    
    def setup_connection_tab(self):
        """Setup database connection tab"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="Database Connection")
        
        # Connection form
        form_frame = ttk.LabelFrame(frame, text="Database Settings", padding=10)
        form_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Label(form_frame, text="Server:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.server_entry = ttk.Entry(form_frame, width=40)
        self.server_entry.grid(row=0, column=1, padx=5, pady=5)
        
        ttk.Label(form_frame, text="Database:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.database_entry = ttk.Entry(form_frame, width=40)
        self.database_entry.grid(row=1, column=1, padx=5, pady=5)
        
        ttk.Label(form_frame, text="Username:").grid(row=2, column=0, sticky=tk.W, pady=5)
        self.username_entry = ttk.Entry(form_frame, width=40)
        self.username_entry.grid(row=2, column=1, padx=5, pady=5)
        
        ttk.Label(form_frame, text="Password:").grid(row=3, column=0, sticky=tk.W, pady=5)
        self.password_entry = ttk.Entry(form_frame, width=40, show="*")
        self.password_entry.grid(row=3, column=1, padx=5, pady=5)
        
        # Connection buttons
        button_frame = ttk.Frame(form_frame)
        button_frame.grid(row=4, column=0, columnspan=2, pady=10)
        
        self.connect_btn = ttk.Button(button_frame, text="Connect", command=self.connect_database)
        self.connect_btn.pack(side=tk.LEFT, padx=5)
        
        self.test_btn = ttk.Button(button_frame, text="Test Connection", command=self.test_connection)
        self.test_btn.pack(side=tk.LEFT, padx=5)
        
        # Connection status
        self.connection_status = ttk.Label(form_frame, text="Not connected", foreground="red")
        self.connection_status.grid(row=5, column=0, columnspan=2, pady=5)
    
    def setup_table_tab(self):
        """Setup table selection tab"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="Table Selection")
        
        # Available tables
        available_frame = ttk.LabelFrame(frame, text="Available Tables", padding=10)
        available_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Treeview for tables
        self.table_tree = ttk.Treeview(available_frame, columns=("Rows", "Columns"), show="tree headings", height=15)
        self.table_tree.heading("#0", text="Table Name")
        self.table_tree.heading("Rows", text="Rows")
        self.table_tree.heading("Columns", text="Columns")
        
        self.table_tree.column("#0", width=200)
        self.table_tree.column("Rows", width=80)
        self.table_tree.column("Columns", width=80)
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(available_frame, orient=tk.VERTICAL, command=self.table_tree.yview)
        self.table_tree.configure(yscrollcommand=scrollbar.set)
        
        self.table_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Selected tables
        selected_frame = ttk.LabelFrame(frame, text="Selected Tables", padding=10)
        selected_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Listbox for selected tables
        self.selected_listbox = tk.Listbox(selected_frame, height=15)
        scrollbar2 = ttk.Scrollbar(selected_frame, orient=tk.VERTICAL, command=self.selected_listbox.yview)
        self.selected_listbox.configure(yscrollcommand=scrollbar2.set)
        
        self.selected_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar2.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Control buttons
        control_frame = ttk.Frame(available_frame)
        control_frame.pack(fill=tk.X, pady=5)
        
        ttk.Button(control_frame, text="Refresh Tables", command=self.refresh_tables).pack(side=tk.LEFT, padx=5)
        ttk.Button(control_frame, text="Add Selected", command=self.add_selected_table).pack(side=tk.LEFT, padx=5)
        ttk.Button(control_frame, text="Remove Selected", command=self.remove_selected_table).pack(side=tk.LEFT, padx=5)
        
        # Fetch button
        fetch_frame = ttk.Frame(frame)
        fetch_frame.pack(fill=tk.X, padx=10, pady=5)
        
        self.fetch_btn = ttk.Button(fetch_frame, text="Fetch Selected Tables", command=self.fetch_selected_tables, state=tk.DISABLED)
        self.fetch_btn.pack(side=tk.RIGHT, padx=5)
    
    def setup_template_tab(self):
        """Setup template configuration tab"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="Template")
        
        # Template file selection
        file_frame = ttk.LabelFrame(frame, text="Template File", padding=10)
        file_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Label(file_frame, text="Template Path:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.template_entry = ttk.Entry(file_frame, width=40)
        self.template_entry.grid(row=0, column=1, padx=5, pady=5)
        
        ttk.Button(file_frame, text="Browse...", command=self.browse_template).grid(row=0, column=2, padx=5)
        
        # Template analysis
        analysis_frame = ttk.LabelFrame(frame, text="Template Analysis", padding=10)
        analysis_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        self.analyze_btn = ttk.Button(analysis_frame, text="Analyze Template", command=self.analyze_template)
        self.analyze_btn.pack(anchor=tk.W, pady=5)
        
        # Analysis results
        self.analysis_text = scrolledtext.ScrolledText(analysis_frame, height=10)
        self.analysis_text.pack(fill=tk.BOTH, expand=True)
    
    def setup_export_tab(self):
        """Setup export configuration tab"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="Export")
        
        # Export settings
        settings_frame = ttk.LabelFrame(frame, text="Export Settings", padding=10)
        settings_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Label(settings_frame, text="Output File:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.output_entry = ttk.Entry(settings_frame, width=40)
        self.output_entry.grid(row=0, column=1, padx=5, pady=5)
        
        ttk.Button(settings_frame, text="Browse...", command=self.browse_output).grid(row=0, column=2, padx=5)
        
        # Auto-generate filename
        self.auto_name_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(settings_frame, text="Auto-generate filename", variable=self.auto_name_var).grid(row=1, column=0, columnspan=2, sticky=tk.W, pady=5)
        
        # Export options
        options_frame = ttk.LabelFrame(frame, text="Export Options", padding=10)
        options_frame.pack(fill=tk.X, padx=10, pady=5)
        
        self.use_template_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(options_frame, text="Use Template", variable=self.use_template_var,
                       command=self.toggle_template_options).grid(row=0, column=0, sticky=tk.W, pady=5)
        
        # Export button
        button_frame = ttk.Frame(frame)
        button_frame.pack(fill=tk.X, padx=10, pady=10)
        
        self.export_btn = ttk.Button(button_frame, text="Export to Excel", command=self.export_data, state=tk.DISABLED)
        self.export_btn.pack(side=tk.RIGHT, padx=5)
        
        # Preview button
        self.preview_btn = ttk.Button(button_frame, text="Preview Data", command=self.preview_data, state=tk.DISABLED)
        self.preview_btn.pack(side=tk.RIGHT, padx=5)
    
    def setup_log_tab(self):
        """Setup log viewer tab"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="Log")
        
        self.log_text = scrolledtext.ScrolledText(frame, wrap=tk.WORD, height=20)
        self.log_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Clear log button
        ttk.Button(frame, text="Clear Log", command=self.clear_log).pack(anchor=tk.E, padx=5, pady=5)
        
        # Redirect logging to the text widget
        class TextHandler(logging.Handler):
            def __init__(self, text_widget):
                super().__init__()
                self.text_widget = text_widget
                
            def emit(self, record):
                msg = self.format(record)
                self.text_widget.insert(tk.END, msg + '\n')
                self.text_widget.see(tk.END)
        
        text_handler = TextHandler(self.log_text)
        text_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
        logger.addHandler(text_handler)
    
    # Event handlers
    def connect_database(self):
        """Connect to database"""
        server = self.server_entry.get()
        database = self.database_entry.get()
        username = self.username_entry.get()
        password = self.password_entry.get()
        
        if not all([server, database, username, password]):
            messagebox.showerror("Error", "Please fill all database connection fields")
            return
        
        self.status_var.set("Connecting to database...")
        self.connect_btn.config(state=tk.DISABLED)
        
        # Run in thread to avoid freezing UI
        def connect_thread():
            success = self.exporter.connect_to_database(server, database, username, password)
            
            self.root.after(0, lambda: self._on_connection_result(success))
        
        threading.Thread(target=connect_thread, daemon=True).start()
    
    def _on_connection_result(self, success: bool):
        """Handle connection result"""
        if success:
            self.connection_status.config(text="Connected", foreground="green")
            self.status_var.set("Connected to database")
            self.refresh_tables()
        else:
            self.connection_status.config(text="Connection failed", foreground="red")
            self.status_var.set("Connection failed")
        
        self.connect_btn.config(state=tk.NORMAL)
    
    def test_connection(self):
        """Test database connection"""
        self.connect_database()
    
    def refresh_tables(self):
        """Refresh list of available tables"""
        if not self.exporter.db_manager or not self.exporter.db_manager.is_connected:
            messagebox.showwarning("Warning", "Not connected to database")
            return
        
        self.status_var.set("Refreshing table list...")
        
        # Clear existing items
        for item in self.table_tree.get_children():
            self.table_tree.delete(item)
        
        # Get table names
        tables = self.exporter.db_manager.get_table_names()
        
        # Add tables to treeview (without row/column counts initially)
        for table in tables:
            self.table_tree.insert("", tk.END, text=table, values=("?", "?"))
        
        self.status_var.set(f"Found {len(tables)} tables")
    
    def add_selected_table(self):
        """Add selected table to list"""
        selection = self.table_tree.selection()
        if not selection:
            return
        
        for item in selection:
            table_name = self.table_tree.item(item, "text")
            if table_name not in self.selected_listbox.get(0, tk.END):
                self.selected_listbox.insert(tk.END, table_name)
        
        self.fetch_btn.config(state=tk.NORMAL)
    
    def remove_selected_table(self):
        """Remove selected table from list"""
        selection = self.selected_listbox.curselection()
        if not selection:
            return
        
        # Remove in reverse order to maintain correct indices
        for index in reversed(selection):
            self.selected_listbox.delete(index)
        
        if self.selected_listbox.size() == 0:
            self.fetch_btn.config(state=tk.DISABLED)
    
    def fetch_selected_tables(self):
        """Fetch data from selected tables"""
        tables = list(self.selected_listbox.get(0, tk.END))
        if not tables:
            messagebox.showwarning("Warning", "No tables selected")
            return
        
        self.status_var.set(f"Fetching {len(tables)} tables...")
        self.fetch_btn.config(state=tk.DISABLED)
        
        def fetch_thread():
            self.exporter.fetch_tables(tables)
            self.root.after(0, self._on_fetch_complete)
        
        threading.Thread(target=fetch_thread, daemon=True).start()
    
    def _on_fetch_complete(self):
        """Handle fetch completion"""
        self.status_var.set("Fetch complete")
        self.fetch_btn.config(state=tk.NORMAL)
        self.export_btn.config(state=tk.NORMAL)
        self.preview_btn.config(state=tk.NORMAL)
        
        # Update table info in treeview
        for table_name, df in self.exporter.tables_data.items():
            for item in self.table_tree.get_children():
                if self.table_tree.item(item, "text") == table_name:
                    self.table_tree.set(item, "Rows", len(df))
                    self.table_tree.set(item, "Columns", len(df.columns))
                    break
    
    def browse_template(self):
        """Browse for template file"""
        filename = filedialog.askopenfilename(
            title="Select Template File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if filename:
            self.template_entry.delete(0, tk.END)
            self.template_entry.insert(0, filename)
    
    def browse_output(self):
        """Browse for output file location"""
        filename = filedialog.asksaveasfilename(
            title="Save As",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if filename:
            self.output_entry.delete(0, tk.END)
            self.output_entry.insert(0, filename)
            self.auto_name_var.set(False)
    
    def analyze_template(self):
        """Analyze selected template"""
        template_path = self.template_entry.get()
        if not template_path or not os.path.exists(template_path):
            messagebox.showerror("Error", "Template file not found")
            return
        
        self.status_var.set("Analyzing template...")
        self.analyze_btn.config(state=tk.DISABLED)
        
        def analyze_thread():
            analysis = self.exporter.analyze_template(template_path)
            self.root.after(0, lambda: self._on_analysis_complete(analysis))
        
        threading.Thread(target=analyze_thread, daemon=True).start()
    
    def _on_analysis_complete(self, analysis: Dict[str, Dict[str, Any]]):
        """Handle template analysis completion"""
        self.analyze_btn.config(state=tk.NORMAL)
        
        self.analysis_text.delete(1.0, tk.END)
        
        if not analysis:
            self.analysis_text.insert(tk.END, "No headers detected in template")
            return
        
        self.analysis_text.insert(tk.END, f"Template Analysis Results:\n")
        self.analysis_text.insert(tk.END, f"Found {len(analysis)} sheet(s) with headers:\n\n")
        
        for sheet_name, info in analysis.items():
            self.analysis_text.insert(tk.END, f"Sheet: {sheet_name}\n")
            self.analysis_text.insert(tk.END, f"  Header Row: {info.get('header_row')}\n")
            self.analysis_text.insert(tk.END, f"  Headers: {len(info.get('headers', []))}\n")
            
            headers = info.get('headers', [])
            for i, header in enumerate(headers[:10]):  # Show first 10 headers
                self.analysis_text.insert(tk.END, f"    {i+1}. {header}\n")
            
            if len(headers) > 10:
                self.analysis_text.insert(tk.END, f"    ... and {len(headers) - 10} more\n")
            
            self.analysis_text.insert(tk.END, "\n")
        
        self.status_var.set("Template analysis complete")
    
    def toggle_template_options(self):
        """Toggle template-related options"""
        use_template = self.use_template_var.get()
        if use_template:
            self.notebook.select(2)  # Switch to template tab
    
    def export_data(self):
        """Export data to Excel"""
        output_path = self.output_entry.get()
        
        # Auto-generate filename if enabled
        if self.auto_name_var.get() or not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"export_{timestamp}.xlsx"
            self.output_entry.delete(0, tk.END)
            self.output_entry.insert(0, output_path)
        
        # Check if we should use template
        template_path = None
        if self.use_template_var.get():
            template_path = self.template_entry.get()
            if not template_path or not os.path.exists(template_path):
                messagebox.showwarning("Warning", "Template file not found. Exporting without template.")
                template_path = None
        
        self.status_var.set("Exporting data...")
        self.export_btn.config(state=tk.DISABLED)
        
        def export_thread():
            success = self.exporter.export_tables(output_path, template_path)
            self.root.after(0, lambda: self._on_export_complete(success, output_path))
        
        threading.Thread(target=export_thread, daemon=True).start()
    
    def _on_export_complete(self, success: bool, output_path: str):
        """Handle export completion"""
        self.export_btn.config(state=tk.NORMAL)
        
        if success:
            self.status_var.set(f"Export complete: {output_path}")
            
            # Ask if user wants to open the file
            response = messagebox.askyesno("Success", 
                                         f"Export completed successfully!\n\nFile saved to:\n{output_path}\n\nOpen file now?")
            if response:
                self.open_file(output_path)
        else:
            self.status_var.set("Export failed")
            messagebox.showerror("Error", "Export failed. Check log for details.")
    
    def preview_data(self):
        """Preview fetched data"""
        if not self.exporter.tables_data:
            messagebox.showinfo("Info", "No data to preview. Fetch tables first.")
            return
        
        # Create preview window
        preview_window = tk.Toplevel(self.root)
        preview_window.title("Data Preview")
        preview_window.geometry("800x600")
        
        # Create notebook for tables
        notebook = ttk.Notebook(preview_window)
        notebook.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        for table_name, df in self.exporter.tables_data.items():
            frame = ttk.Frame(notebook)
            notebook.add(frame, text=f"{table_name} ({len(df)} rows)")
            
            # Create treeview for this table
            tree = ttk.Treeview(frame, columns=list(df.columns), show="headings")
            
            # Configure columns
            for col in df.columns:
                tree.heading(col, text=col)
                tree.column(col, width=100)
            
            # Add data (limit to 100 rows for performance)
            for i, row in df.head(100).iterrows():
                tree.insert("", tk.END, values=list(row))
            
            # Add scrollbar
            scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=tree.yview)
            tree.configure(yscrollcommand=scrollbar.set)
            
            tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            # Show row count
            ttk.Label(frame, text=f"Showing {min(100, len(df))} of {len(df)} rows").pack(side=tk.BOTTOM, pady=5)
    
    def open_file(self, filepath: str):
        """Open a file with default application"""
        try:
            if platform.system() == "Windows":
                os.startfile(filepath)
            elif platform.system() == "Darwin":  # macOS
                subprocess.run(["open", filepath])
            else:  # Linux
                subprocess.run(["xdg-open", filepath])
        except Exception as e:
            messagebox.showerror("Error", f"Could not open file: {str(e)}")
    
    def clear_log(self):
        """Clear log text widget"""
        self.log_text.delete(1.0, tk.END)

# ============================================================================
# DATABASE MANAGER (FIXED)
# ============================================================================

class DatabaseManager:
    """Manages database connections and queries"""
    
    def __init__(self, connection_string: Optional[str] = None):
        self.connection_string = connection_string
        self.connection = None
        self.is_connected = False
    
    def connect(self, server: str, database: str, username: str, password: str,
                driver: str = "ODBC Driver 17 for SQL Server") -> bool:
        """Connect to SQL Server database"""
        try:
            self.connection_string = (
                f"DRIVER={{{driver}}};"
                f"SERVER={server};"
                f"DATABASE={database};"
                f"UID={username};"
                f"PWD={password};"
                "TrustServerCertificate=yes;"
            )
            self.connection = pyodbc.connect(self.connection_string)
            self.is_connected = True
            logger.info(f"✅ Connected to database: {database} on {server}")
            return True
        except Exception as e:
            logger.error(f"❌ Database connection failed: {e}")
            self.is_connected = False
            return False
    
    def get_table_names(self) -> List[str]:
        """Get list of table names in the database"""
        if not self.is_connected:
            return []
        
        try:
            cursor = self.connection.cursor()
            cursor.execute("""
                SELECT TABLE_NAME 
                FROM INFORMATION_SCHEMA.TABLES 
                WHERE TABLE_TYPE = 'BASE TABLE'
                ORDER BY TABLE_NAME
            """)
            tables = [row[0] for row in cursor.fetchall()]
            return tables
        except Exception as e:
            logger.error(f"❌ Failed to get table names: {e}")
            return []
    
    def get_column_names(self, table_name: str) -> List[str]:
        """Get column names for a specific table"""
        if not self.is_connected:
            return []
        
        try:
            cursor = self.connection.cursor()
            cursor.execute(f"""
                SELECT COLUMN_NAME 
                FROM INFORMATION_SCHEMA.COLUMNS 
                WHERE TABLE_NAME = '{table_name}'
                ORDER BY ORDINAL_POSITION
            """)
            columns = [row[0] for row in cursor.fetchall()]
            return columns
        except Exception as e:
            logger.error(f"❌ Failed to get columns for table {table_name}: {e}")
            return []
    
    def query_data(self, query: str, params: Optional[tuple] = None) -> List[tuple]:
        """Execute a query and return results"""
        if not self.is_connected:
            return []
        
        try:
            cursor = self.connection.cursor()
            if params:
                cursor.execute(query, params)
            else:
                cursor.execute(query)
            return cursor.fetchall()
        except Exception as e:
            logger.error(f"❌ Query failed: {e}")
            return []
    
    def close(self):
        """Close database connection"""
        if self.connection:
            self.connection.close()
            self.is_connected = False
            logger.info("Database connection closed")

# ============================================================================
# MAIN FUNCTION
# ============================================================================

def main():
    """Main entry point"""
    # Check for required packages
    required_packages = ['pyodbc', 'openpyxl', 'PIL', 'tkcalendar']
    missing_packages = []
    
    for package in required_packages:
        try:
            __import__(package.replace('-', '_'))
        except ImportError:
            missing_packages.append(package)
    
    if missing_packages:
        print("Missing required packages:")
        for pkg in missing_packages:
            print(f"  - {pkg}")
        print("\nInstall with: pip install " + " ".join(missing_packages))
        return
    
    # Create and run GUI
    root = tk.Tk()
    
    # Set window icon if available
    try:
        root.iconbitmap(default='icon.ico')
    except:
        pass
    
    app = TableExporterGUI(root)
    
    # Center the window
    root.update_idletasks()
    width = root.winfo_width()
    height = root.winfo_height()
    x = (root.winfo_screenwidth() // 2) - (width // 2)
    y = (root.winfo_screenheight() // 2) - (height // 2)
    root.geometry(f'{width}x{height}+{x}+{y}')
    
    # Start the application
    try:
        root.mainloop()
    except KeyboardInterrupt:
        print("\nApplication terminated by user")
    except Exception as e:
        print(f"Application error: {e}")
        logger.error(f"Application error: {e}", exc_info=True)

if __name__ == "__main__":
    main()