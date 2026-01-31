import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext, filedialog
import pyodbc
import json
import os
import logging
from datetime import datetime
import threading
import sys
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
import traceback
import shutil
from typing import Dict, List, Optional, Any, Tuple
from dataclasses import dataclass, field
import re

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('table_exporter.log'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# ============================================================================
# DATABASE MANAGER
# ============================================================================

class DatabaseManager:
    """Manages database connections and queries"""
    
    def __init__(self):
        self.connection = None
        self.connected = False
        self.server = None
        self.database = None
    
    def connect(self, server: str, database: str, 
                username: str = None, password: str = None,
                use_windows_auth: bool = True) -> Tuple[bool, str]:
        """Connect to SQL Server database"""
        try:
            # Build connection string
            if use_windows_auth:
                conn_str = f"DRIVER={{SQL Server}};SERVER={server};DATABASE={database};Trusted_Connection=yes;"
            else:
                conn_str = f"DRIVER={{SQL Server}};SERVER={server};DATABASE={database};UID={username};PWD={password};"
            
            # Attempt connection
            self.connection = pyodbc.connect(conn_str)
            self.connected = True
            self.server = server
            self.database = database
            
            logger.info(f"Connected to {server}.{database}")
            return True, "Connection successful"
            
        except pyodbc.Error as e:
            error_msg = f"Database connection error: {str(e)}"
            logger.error(error_msg)
            self.connected = False
            return False, error_msg
        except Exception as e:
            error_msg = f"Unexpected error during connection: {str(e)}"
            logger.error(error_msg)
            self.connected = False
            return False, error_msg
    
    def disconnect(self):
        """Disconnect from database"""
        try:
            if self.connection:
                self.connection.close()
                logger.info(f"Disconnected from {self.server}.{self.database}")
        except Exception as e:
            logger.error(f"Error disconnecting: {e}")
        finally:
            self.connection = None
            self.connected = False
            self.server = None
            self.database = None
    
    def get_tables(self) -> List[str]:
        """Get list of tables in the database"""
        try:
            cursor = self.connection.cursor()
            
            # Query to get user tables (excluding system tables)
            query = """
            SELECT TABLE_NAME 
            FROM INFORMATION_SCHEMA.TABLES 
            WHERE TABLE_TYPE = 'BASE TABLE'
            ORDER BY TABLE_NAME
            """
            
            cursor.execute(query)
            tables = [row[0] for row in cursor.fetchall()]
            cursor.close()
            
            logger.info(f"Retrieved {len(tables)} tables")
            return tables
            
        except Exception as e:
            logger.error(f"Error getting tables: {e}")
            raise
    
    def get_table_columns(self, table_name: str) -> List[str]:
        """Get column names for a specific table"""
        try:
            cursor = self.connection.cursor()
            
            # Query to get column names
            query = f"""
            SELECT COLUMN_NAME 
            FROM INFORMATION_SCHEMA.COLUMNS 
            WHERE TABLE_NAME = ?
            ORDER BY ORDINAL_POSITION
            """
            
            cursor.execute(query, (table_name,))
            columns = [row[0] for row in cursor.fetchall()]
            cursor.close()
            
            logger.info(f"Retrieved {len(columns)} columns for table {table_name}")
            return columns
            
        except Exception as e:
            logger.error(f"Error getting columns for {table_name}: {e}")
            raise
    
    def fetch_table_data(self, table_name: str, limit: int = None) -> Dict:
        """Fetch data from a specific table"""
        try:
            # Get display name
            display_name = self.get_display_name(table_name)
            
            # Get columns
            columns = self.get_table_columns(table_name)
            
            if not columns:
                return {
                    'success': False,
                    'error': f"No columns found for table: {table_name}",
                    'display_name': display_name,
                    'table_name': table_name
                }
            
            # Build query
            column_list = ', '.join([f'[{col}]' for col in columns])
            query = f"SELECT {column_list} FROM [{table_name}]"
            
            if limit and limit > 0:
                query += f" ORDER BY (SELECT NULL) OFFSET 0 ROWS FETCH NEXT {limit} ROWS ONLY"
            
            # Execute query
            cursor = self.connection.cursor()
            cursor.execute(query)
            
            # Fetch all rows
            rows = cursor.fetchall()
            row_count = len(rows)
            
            # Convert to list of dictionaries
            data = []
            for row in rows:
                row_dict = {}
                for i, col in enumerate(columns):
                    row_dict[col] = row[i]
                data.append(row_dict)
            
            cursor.close()
            
            logger.info(f"Fetched {row_count} rows from {table_name}")
            
            return {
                'success': True,
                'display_name': display_name,
                'table_name': table_name,
                'columns': columns,
                'data': data,
                'row_count': row_count
            }
            
        except Exception as e:
            error_msg = f"Error fetching data from {table_name}: {str(e)}"
            logger.error(error_msg)
            return {
                'success': False,
                'error': error_msg,
                'display_name': self.get_display_name(table_name),
                'table_name': table_name,
                'columns': [],
                'data': [],
                'row_count': 0
            }
    
    def get_display_name(self, table_name: str) -> str:
        """Convert table name to display name"""
        # Remove underscores and capitalize
        display_name = table_name.replace('_', ' ').title()
        
        # Common replacements
        replacements = {
            'Tbl': '',
            'Table': '',
            'Vw': 'View: ',
            'Vw_': 'View: ',
            'View': 'View: '
        }
        
        for old, new in replacements.items():
            if display_name.startswith(old):
                display_name = display_name.replace(old, new, 1).strip()
        
        return display_name if display_name else table_name

# ============================================================================
# DATA CLASSES
# ============================================================================

@dataclass
class CellMapping:
    """Mapping information for a single cell"""
    table_name: str
    column_name: str
    template_sheet: str
    template_cell: str  # e.g., "A1", "B3", "C5"
    apply_to_all_sheets: bool = False
    selected_sheets: List[str] = field(default_factory=list)  # NEW: Specific sheets
    
@dataclass
class TableConfig:
    """Configuration for a data table"""
    table_name: str
    display_name: str
    start_row: int
    start_col: str
    sheet_name: str
    column_mappings: Dict[str, CellMapping] = field(default_factory=dict)
    apply_to_all_sheets: bool = False
    selected_sheets: List[str] = field(default_factory=list)  # NEW: Specific sheets

# ============================================================================
# SHEET SELECTION DIALOG
# ============================================================================

class SheetSelectionDialog:
    """Dialog for selecting specific sheets"""
    
    def __init__(self, parent, available_sheets: List[str], title: str = "Select Sheets"):
        self.parent = parent
        self.available_sheets = available_sheets
        self.title = title
        self.selected_sheets = []
        self.result = None
        
        self.create_dialog()
    
    def create_dialog(self):
        """Create the sheet selection dialog"""
        self.dialog = tk.Toplevel(self.parent)
        self.dialog.title(self.title)
        self.dialog.geometry("400x500")
        self.dialog.transient(self.parent)
        self.dialog.grab_set()
        
        # Center dialog
        self.dialog.update_idletasks()
        x = (self.dialog.winfo_screenwidth() - self.dialog.winfo_width()) // 2
        y = (self.dialog.winfo_screenheight() - self.dialog.winfo_height()) // 2
        self.dialog.geometry(f"+{x}+{y}")
        
        # Main container
        main_frame = ttk.Frame(self.dialog, padding="20")
        main_frame.pack(fill='both', expand=True)
        
        # Title
        ttk.Label(main_frame, text="Select Sheets", 
                 font=('Arial', 14, 'bold')).pack(pady=(0, 10))
        ttk.Label(main_frame, text=f"Available sheets: {len(self.available_sheets)}", 
                 font=('Arial', 10)).pack(pady=(0, 20))
        
        # Instructions
        instr_frame = ttk.LabelFrame(main_frame, text="Instructions", padding="10")
        instr_frame.pack(fill='x', pady=(0, 20))
        
        ttk.Label(instr_frame, text="Select which sheets to apply the data to:", 
                 font=('Arial', 9)).pack(anchor='w', pady=2)
        ttk.Label(instr_frame, text="• Check sheets where data should appear", 
                 font=('Arial', 9)).pack(anchor='w', pady=1)
        ttk.Label(instr_frame, text="• Leave unchecked to skip", 
                 font=('Arial', 9)).pack(anchor='w', pady=1)
        
        # Create scrollable frame for sheet checkboxes
        canvas = tk.Canvas(main_frame, highlightthickness=0)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        sheet_frame = ttk.Frame(canvas)
        
        canvas.pack(side=tk.LEFT, fill='both', expand=True)
        scrollbar.pack(side=tk.RIGHT, fill='y')
        
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.create_window((0, 0), window=sheet_frame, anchor="nw")
        
        # Store variables
        self.sheet_vars = {}
        
        # Create checkboxes for each sheet
        for sheet in self.available_sheets:
            var = tk.BooleanVar(value=True)  # Default selected
            self.sheet_vars[sheet] = var
            
            cb_frame = ttk.Frame(sheet_frame)
            cb_frame.pack(fill='x', padx=5, pady=2)
            
            cb = ttk.Checkbutton(cb_frame, text=sheet, variable=var)
            cb.pack(anchor='w')
        
        # Configure scrolling
        def on_frame_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
        
        sheet_frame.bind("<Configure>", on_frame_configure)
        
        # Control buttons
        control_frame = ttk.Frame(sheet_frame)
        control_frame.pack(fill='x', pady=10)
        
        ttk.Button(control_frame, text="Select All", 
                  command=self.select_all).pack(side=tk.LEFT, padx=5)
        ttk.Button(control_frame, text="Clear All", 
                  command=self.clear_all).pack(side=tk.LEFT, padx=5)
        
        # Buttons
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill='x', pady=20)
        
        ttk.Button(btn_frame, text="Apply", command=self.apply, 
                  style='Accent.TButton').pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Cancel", command=self.cancel).pack(side=tk.RIGHT, padx=5)
        
        # Clean up on close
        self.dialog.protocol("WM_DELETE_WINDOW", self.cancel)
    
    def select_all(self):
        """Select all sheets"""
        for var in self.sheet_vars.values():
            var.set(True)
    
    def clear_all(self):
        """Clear all sheet selections"""
        for var in self.sheet_vars.values():
            var.set(False)
    
    def apply(self):
        """Apply sheet selection"""
        self.selected_sheets = [sheet for sheet, var in self.sheet_vars.items() if var.get()]
        
        if not self.selected_sheets:
            messagebox.showwarning("No Selection", "Please select at least one sheet.")
            return
        
        self.result = self.selected_sheets
        self.dialog.destroy()
    
    def cancel(self):
        """Cancel sheet selection"""
        self.result = None
        self.dialog.destroy()
    
    def get_selected_sheets(self) -> Optional[List[str]]:
        """Get the selection result"""
        return self.result

# ============================================================================
# POSITION MAPPING DIALOG WITH SHEET SELECTION
# ============================================================================

class PositionMappingDialog:
    """Dialog for mapping database columns to template cells"""
    
    def __init__(self, parent, table_name: str, db_columns: List[str], 
                 template_sheets: List[str]):
        self.parent = parent
        self.table_name = table_name
        self.db_columns = db_columns
        self.template_sheets = template_sheets
        self.mappings = {}  # column_name -> (sheet_name, cell_reference, apply_to_all, selected_sheets)
        self.result = None
        
        self.create_dialog()
    
    def create_dialog(self):
        """Create the position mapping dialog"""
        self.dialog = tk.Toplevel(self.parent)
        self.dialog.title(f"Map Positions for {self.table_name}")
        self.dialog.geometry("1000x600")
        self.dialog.transient(self.parent)
        self.dialog.grab_set()
        
        # Center dialog
        self.dialog.update_idletasks()
        x = (self.dialog.winfo_screenwidth() - self.dialog.winfo_width()) // 2
        y = (self.dialog.winfo_screenheight() - self.dialog.winfo_height()) // 2
        self.dialog.geometry(f"+{x}+{y}")
        
        # Main container
        main_frame = ttk.Frame(self.dialog, padding="10")
        main_frame.pack(fill='both', expand=True)
        
        # Title
        ttk.Label(main_frame, text=f"Map Database Columns to Template Cells", 
                 font=('Arial', 14, 'bold')).pack(pady=(0, 10))
        ttk.Label(main_frame, text=f"Table: {self.table_name}", 
                 font=('Arial', 11)).pack(pady=(0, 20))
        
        # Instructions
        instr_frame = ttk.LabelFrame(main_frame, text="Instructions", padding="10")
        instr_frame.pack(fill='x', pady=(0, 20))
        
        instructions = [
            "1. For each database column, select:",
            "   • Sheet: Choose base sheet or 'All Sheets'",
            "   • Cell: Enter cell reference (e.g., B4, C4, D4)",
            "   • Apply to: Choose 'All Sheets', 'This Sheet Only', or 'Select Sheets'",
            "2. Leave cell empty to skip that column",
            "3. Example: BATCH_NAME → Sheet1 → B4 → All Sheets",
            "4. Example: JOB_NO → Sheet1 → C4 → Select Sheets (Sheet1, Sheet3)",
            "5. For merged cells: Write to top-left cell (e.g., B4 for B4:D4 merged)"
        ]
        
        for instr in instructions:
            ttk.Label(instr_frame, text=instr, font=('Arial', 9)).pack(anchor='w', pady=1)
        
        # Create scrollable frame for mappings
        canvas = tk.Canvas(main_frame, highlightthickness=0)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        mapping_frame = ttk.Frame(canvas)
        
        canvas.pack(side=tk.LEFT, fill='both', expand=True)
        scrollbar.pack(side=tk.RIGHT, fill='y')
        
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.create_window((0, 0), window=mapping_frame, anchor="nw")
        
        # Store variables
        self.sheet_vars = {}
        self.cell_vars = {}
        self.preview_vars = {}
        self.apply_type_vars = {}  # "all", "this", "select"
        self.selected_sheets_vars = {}  # For "select" option
        
        # Create mapping rows
        header_frame = ttk.Frame(mapping_frame)
        header_frame.pack(fill='x', pady=(0, 10))
        
        columns = [
            ("Database Column", 20),
            ("Sheet", 15),
            ("Cell", 10),
            ("Apply to", 15),
            ("Sheets Selected", 20),
            ("Preview", 25)
        ]
        
        for i, (text, width) in enumerate(columns):
            ttk.Label(header_frame, text=text, width=width, 
                     font=('Arial', 10, 'bold')).grid(row=0, column=i, padx=5)
        
        # Create rows for each database column
        for i, column in enumerate(self.db_columns, 1):
            row_frame = ttk.Frame(mapping_frame)
            row_frame.pack(fill='x', pady=2)
            
            # Database column name
            ttk.Label(row_frame, text=column, width=20).grid(row=0, column=0, padx=5)
            
            # Sheet dropdown
            sheet_var = tk.StringVar(value=self.template_sheets[0] if self.template_sheets else "")
            sheet_cb = ttk.Combobox(row_frame, textvariable=sheet_var, 
                                   values=self.template_sheets, width=15, state="readonly")
            sheet_cb.grid(row=0, column=1, padx=5)
            self.sheet_vars[column] = sheet_var
            
            # Cell reference entry
            cell_var = tk.StringVar()
            cell_entry = ttk.Entry(row_frame, textvariable=cell_var, width=10)
            cell_entry.grid(row=0, column=2, padx=5)
            self.cell_vars[column] = cell_var
            
            # Apply type dropdown
            apply_var = tk.StringVar(value="this")  # "all", "this", "select"
            apply_cb = ttk.Combobox(row_frame, textvariable=apply_var, 
                                   values=["All Sheets", "This Sheet Only", "Select Sheets"], 
                                   width=15, state="readonly")
            apply_cb.grid(row=0, column=3, padx=5)
            self.apply_type_vars[column] = apply_var
            
            # Select sheets button (initially disabled)
            select_btn = ttk.Button(row_frame, text="Select Sheets...", 
                                   command=lambda col=column: self.select_sheets(col),
                                   width=15, state='disabled')
            select_btn.grid(row=0, column=4, padx=5)
            self.selected_sheets_vars[column] = {
                'button': select_btn,
                'selected': [self.template_sheets[0]] if self.template_sheets else []
            }
            
            # Preview label
            preview_var = tk.StringVar(value="Not mapped")
            ttk.Label(row_frame, textvariable=preview_var, width=25).grid(row=0, column=5, padx=5)
            self.preview_vars[column] = preview_var
            
            # Add validation and event handlers
            def on_change(col_name, *args):
                self.update_preview(col_name)
                # Enable/disable select sheets button
                apply_type = self.apply_type_vars[col_name].get()
                if apply_type == "Select Sheets":
                    self.selected_sheets_vars[col_name]['button'].config(state='normal')
                else:
                    self.selected_sheets_vars[col_name]['button'].config(state='disabled')
            
            # Use trace_add
            cell_var.trace_add('write', lambda name, index, mode, col=column: on_change(col))
            sheet_var.trace_add('write', lambda name, index, mode, col=column: on_change(col))
            apply_var.trace_add('write', lambda name, index, mode, col=column: on_change(col))
        
        # Configure scrolling
        def on_frame_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
        
        mapping_frame.bind("<Configure>", on_frame_configure)
        
        # Safe mouse wheel binding
        def on_mousewheel(event):
            try:
                canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            except tk.TclError:
                pass
        
        canvas.bind("<MouseWheel>", on_mousewheel)
        
        # Buttons
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill='x', pady=20)
        
        ttk.Button(btn_frame, text="Apply Mappings", command=self.apply_mappings, 
                  style='Accent.TButton').pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Cancel", command=self.cancel).pack(side=tk.RIGHT, padx=5)
        
        # Clean up on close
        self.dialog.protocol("WM_DELETE_WINDOW", self.cancel)
    
    def select_sheets(self, column_name: str):
        """Open sheet selection dialog for a specific column"""
        dialog = SheetSelectionDialog(self.dialog, self.template_sheets, 
                                     f"Select Sheets for {column_name}")
        self.dialog.wait_window(dialog.dialog)
        
        selected = dialog.get_selected_sheets()
        if selected:
            self.selected_sheets_vars[column_name]['selected'] = selected
            self.update_preview(column_name)
    
    def update_preview(self, column_name: str):
        """Update preview for a column"""
        cell_val = self.cell_vars[column_name].get().strip().upper()
        sheet_val = self.sheet_vars[column_name].get()
        apply_type = self.apply_type_vars[column_name].get()
        
        if cell_val:
            if re.match(r'^[A-Z]+\d+$', cell_val):
                if apply_type == "All Sheets":
                    self.preview_vars[column_name].set(f"Will write to {cell_val} on ALL sheets")
                elif apply_type == "This Sheet Only":
                    self.preview_vars[column_name].set(f"Will write to {cell_val} on {sheet_val}")
                elif apply_type == "Select Sheets":
                    selected = self.selected_sheets_vars[column_name]['selected']
                    if selected:
                        sheet_list = ", ".join(selected[:2])
                        if len(selected) > 2:
                            sheet_list += f" (+{len(selected)-2} more)"
                        self.preview_vars[column_name].set(f"Will write to {cell_val} on: {sheet_list}")
                    else:
                        self.preview_vars[column_name].set(f"Will write to {cell_val} (select sheets)")
            else:
                self.preview_vars[column_name].set("Invalid cell format")
        else:
            self.preview_vars[column_name].set("Not mapped")
    
    def apply_mappings(self):
        """Apply all mappings"""
        self.mappings = {}
        
        for column in self.db_columns:
            sheet = self.sheet_vars[column].get()
            cell = self.cell_vars[column].get().strip().upper()
            apply_type = self.apply_type_vars[column].get()
            selected_sheets = self.selected_sheets_vars[column]['selected']
            
            if sheet and cell:
                # Validate cell format
                if re.match(r'^[A-Z]+\d+$', cell):
                    # Determine apply settings
                    apply_all = (apply_type == "All Sheets")
                    
                    # For "Select Sheets", use the selected sheets
                    target_sheets = []
                    if apply_type == "Select Sheets":
                        if not selected_sheets:
                            messagebox.showerror("No Sheets Selected", 
                                               f"No sheets selected for {column}. Please select sheets.")
                            return
                        target_sheets = selected_sheets
                    elif apply_type == "This Sheet Only":
                        target_sheets = [sheet]
                    else:  # All Sheets
                        target_sheets = self.template_sheets
                    
                    self.mappings[column] = (sheet, cell, apply_all, target_sheets)
                else:
                    messagebox.showerror("Invalid Cell", 
                                       f"Invalid cell reference for {column}: {cell}\n\n"
                                       "Use format like: B4, C4, D10, etc.")
                    return
        
        if not self.mappings:
            if messagebox.askyesno("No Mappings", 
                                  "No columns have been mapped. Continue anyway?"):
                self.result = {}
                self.dialog.destroy()
            return
        else:
            self.result = self.mappings
            self.dialog.destroy()
    
    def cancel(self):
        """Cancel mapping"""
        self.result = None
        self.dialog.destroy()
    
    def get_mappings(self) -> Optional[Dict]:
        """Get the mapping result"""
        return self.result

# ============================================================================
# EXCEL EXPORTER WITH SHEET SELECTION SUPPORT
# ============================================================================

class ExcelTableExporter:
    """Handles exporting tables to Excel with position mapping and merged cell support"""
    
    @staticmethod
    def export_tables_to_excel(tables_data: Dict, output_path: str) -> bool:
        """Export multiple tables to new Excel file"""
        try:
            wb = Workbook()
            
            # Remove default sheet
            if wb.sheetnames:
                wb.remove(wb.active)
            
            # Create a sheet for each table
            for table_name, table_data in tables_data.items():
                if table_data.get('success', False):
                    # Create sheet with valid name
                    sheet_name = ExcelTableExporter.get_valid_sheet_name(table_data['display_name'])
                    ws = wb.create_sheet(title=sheet_name)
                    
                    # Add table to sheet
                    ExcelTableExporter.add_table_to_sheet(ws, table_data)
            
            # Save workbook
            wb.save(output_path)
            logger.info(f"Excel file created: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Excel export error: {e}")
            raise
    
    @staticmethod
    def export_tables_to_template(tables_data: Dict, template_path: str, 
                                table_configs: Dict[str, TableConfig],
                                output_path: str,
                                merge_rules: List[str] = None) -> bool:
        """
        Export data into an existing template using position mappings.
        Supports writing to selected sheets.
        """
        try:
            # Make a copy of the template
            shutil.copy2(template_path, output_path)
            
            # Load the copied template
            wb = load_workbook(output_path)
            
            # Apply user merge rules first (optional)
            if merge_rules:
                for rule in merge_rules:
                    try:
                        if "!" in rule:
                            sheet_name, cell_range = rule.split("!", 1)
                            sheet_name = sheet_name.strip()
                            cell_range = cell_range.strip()
                        else:
                            continue
                        if sheet_name in wb.sheetnames:
                            wb[sheet_name].merge_cells(cell_range)
                    except Exception:
                        # ignore invalid merge rules
                        pass
            
            # Process each table
            for table_name, table_data in tables_data.items():
                if not table_data.get('success', False):
                    continue
                
                if table_name not in table_configs:
                    logger.warning(f"No configuration found for table: {table_name}")
                    continue
                
                table_config = table_configs[table_name]
                
                # Write individual column mappings
                for column_name, cell_mapping in table_config.column_mappings.items():
                    # Check if this column exists in the data
                    if column_name in table_data['columns'] and table_data['data']:
                        # Get the value from first row
                        value = table_data['data'][0].get(column_name, "")
                        
                        # Determine which sheets to write to
                        sheets_to_write = []
                        if cell_mapping.apply_to_all_sheets or table_config.apply_to_all_sheets:
                            # Write to all sheets
                            sheets_to_write = wb.sheetnames
                        elif cell_mapping.selected_sheets:
                            # Write to selected sheets
                            sheets_to_write = [s for s in cell_mapping.selected_sheets if s in wb.sheetnames]
                        elif table_config.selected_sheets:
                            # Write to table's selected sheets
                            sheets_to_write = [s for s in table_config.selected_sheets if s in wb.sheetnames]
                        else:
                            # Write to specific sheet only
                            if cell_mapping.template_sheet in wb.sheetnames:
                                sheets_to_write = [cell_mapping.template_sheet]
                        
                        # Write to each sheet
                        for sheet_name in sheets_to_write:
                            success = ExcelTableExporter.write_to_cell_safe(
                                wb, 
                                sheet_name, 
                                cell_mapping.template_cell, 
                                value
                            )
                            
                            if success:
                                logger.info(f"Wrote {column_name}={value} to {sheet_name}!{cell_mapping.template_cell}")
                            else:
                                logger.warning(f"Could not write to {sheet_name}!{cell_mapping.template_cell}")
                
                # Write tabular data if start position is configured
                if table_config.start_row > 0 and table_config.start_col:
                    # Determine which sheets to write to
                    sheets_to_write = []
                    if table_config.apply_to_all_sheets:
                        sheets_to_write = wb.sheetnames
                    elif table_config.selected_sheets:
                        sheets_to_write = [s for s in table_config.selected_sheets if s in wb.sheetnames]
                    else:
                        if table_config.sheet_name in wb.sheetnames:
                            sheets_to_write = [table_config.sheet_name]
                    
                    for sheet_name in sheets_to_write:
                        ws = wb[sheet_name]
                        start_col_idx = column_index_from_string(table_config.start_col)
                        
                        # Find first safe row
                        safe_row = ExcelTableExporter.find_safe_row_for_table(ws, table_config.start_row)
                        
                        # Write headers
                        for col_idx, col_name in enumerate(table_data['columns'], start=0):
                            cell_col = start_col_idx + col_idx
                            cell_ref = f"{get_column_letter(cell_col)}{safe_row}"
                            
                            success = ExcelTableExporter.write_to_cell_safe(
                                wb, sheet_name, cell_ref, col_name
                            )
                            
                            if success:
                                cell = ws[cell_ref]
                                cell.font = Font(bold=True)
                                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                        
                        # Write data
                        for row_idx, row_data in enumerate(table_data['data'], start=1):
                            for col_idx, col_name in enumerate(table_data['columns'], start=0):
                                value = row_data.get(col_name, "")
                                cell_col = start_col_idx + col_idx
                                cell_row = safe_row + row_idx
                                cell_ref = f"{get_column_letter(cell_col)}{cell_row}"
                                
                                ExcelTableExporter.write_to_cell_safe(
                                    wb, sheet_name, cell_ref, value
                                )
            
            # Save workbook
            wb.save(output_path)
            logger.info(f"Template exported: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Template export error: {e}")
            raise
    
    @staticmethod
    def write_to_cell_safe(wb, sheet_name: str, cell_ref: str, value: Any) -> bool:
        """
        Safely write to a cell, handling merged cells.
        Returns True if successful, False if cell is in a merged range.
        """
        try:
            if sheet_name not in wb.sheetnames:
                return False
            
            ws = wb[sheet_name]
            
            # Parse cell reference
            col_letter = ''.join([c for c in cell_ref if c.isalpha()])
            row_num = int(''.join([c for c in cell_ref if c.isdigit()]))
            col_num = column_index_from_string(col_letter)
            
            # Check if cell is part of a merged range
            for merged_range in ws.merged_cells.ranges:
                if (merged_range.min_row <= row_num <= merged_range.max_row and
                    merged_range.min_col <= col_num <= merged_range.max_col):
                    # Cell is in a merged range
                    # Try to write to the top-left cell of the merged range
                    top_left_cell = f"{get_column_letter(merged_range.min_col)}{merged_range.min_row}"
                    try:
                        ws[top_left_cell] = value
                        # Center alignment for merged cells
                        ws[top_left_cell].alignment = Alignment(horizontal='center', vertical='center')
                        return True
                    except:
                        return False
            
            # Cell is not merged, write normally
            ws[cell_ref] = value
            return True
            
        except Exception as e:
            logger.error(f"Error writing to cell {sheet_name}!{cell_ref}: {e}")
            return False
    
    @staticmethod
    def find_safe_row_for_table(ws, start_row: int) -> int:
        """Find a safe row to start table data (not in merged cells)"""
        current_row = start_row
        
        # Check if start row is safe (not in merged cells in first 5 columns)
        for col in range(1, 6):  # Check first 5 columns
            cell_ref = f"{get_column_letter(col)}{current_row}"
            for merged_range in ws.merged_cells.ranges:
                if cell_ref in merged_range:
                    # Row is in merged range, try next row
                    current_row += 1
                    return ExcelTableExporter.find_safe_row_for_table(ws, current_row)
        
        return current_row
    
    @staticmethod
    def get_valid_sheet_name(name: str) -> str:
        """Get valid Excel sheet name"""
        # Remove invalid characters
        invalid_chars = ['\\', '/', '*', '?', ':', '[', ']']
        for char in invalid_chars:
            name = name.replace(char, '_')
        
        # Truncate if too long
        if len(name) > 31:
            name = name[:28] + "..."
        
        # Ensure not empty
        if not name.strip():
            name = "Sheet"
        
        return name[:31]
    
    @staticmethod
    def add_table_to_sheet(ws, table_data: Dict):
        """Add a table to Excel sheet"""
        # Write headers
        for col_idx, col_name in enumerate(table_data['columns'], start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Write data
        for row_idx, row_data in enumerate(table_data['data'], start=2):
            for col_idx, col_name in enumerate(table_data['columns'], start=1):
                value = row_data.get(col_name, '')
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

# ============================================================================
# MAIN APPLICATION WITH SHEET SELECTION
# ============================================================================

class MultiTableExporterApp:
    """Main Application"""
    
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Table Exporter with Sheet Selection")
        self.root.geometry("1200x800")
        
        # Database connection
        self.db = DatabaseManager()
        self.exporter = ExcelTableExporter()
        
        # Variables
        self.server_var = tk.StringVar(value="MAHESHWAGH\\WINCC")
        self.database_var = tk.StringVar(value="VPI1")
        
        # Table selection and configuration
        self.selected_tables = []
        self.table_checkboxes = {}
        self.table_configs = {}  # table_name -> TableConfig
        
        # Template
        self.template_path = None
        self.template_sheets = []
        
        # Merge rules
        self.merge_rules = []
        
        # Setup UI
        self.setup_ui()
        
        # Bind window close
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
    
    def setup_ui(self):
        """Setup the user interface"""
        # Create main container
        main_container = ttk.Frame(self.root, padding="10")
        main_container.pack(fill='both', expand=True)
        
        # Create notebook for tabs
        self.notebook = ttk.Notebook(main_container)
        self.notebook.pack(fill='both', expand=True)
        
        # Create tabs
        self.setup_connection_tab()
        self.setup_table_selection_tab()
        self.setup_position_mapping_tab()
        self.setup_export_tab()
        
        # Status bar
        self.status_bar = ttk.Label(self.root, text="Ready", relief=tk.SUNKEN, anchor=tk.W)
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)
    
    def setup_connection_tab(self):
        """Setup Connection Tab"""
        conn_tab = ttk.Frame(self.notebook)
        self.notebook.add(conn_tab, text="Connection")
        
        # Main frame
        main_frame = ttk.Frame(conn_tab, padding="20")
        main_frame.pack(fill='both', expand=True)
        
        # Title
        ttk.Label(main_frame, text="Database Connection", font=('Arial', 16, 'bold')).pack(pady=(0, 20))
        
        # Server
        server_frame = ttk.Frame(main_frame)
        server_frame.pack(fill='x', pady=5)
        ttk.Label(server_frame, text="Server:", width=15).pack(side=tk.LEFT)
        ttk.Entry(server_frame, textvariable=self.server_var).pack(side=tk.LEFT, fill='x', expand=True, padx=5)
        
        # Database
        db_frame = ttk.Frame(main_frame)
        db_frame.pack(fill='x', pady=5)
        ttk.Label(db_frame, text="Database:", width=15).pack(side=tk.LEFT)
        ttk.Entry(db_frame, textvariable=self.database_var).pack(side=tk.LEFT, fill='x', expand=True, padx=5)
        
        # Connection buttons
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(pady=20)
        ttk.Button(btn_frame, text="Connect", command=self.connect_db).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Disconnect", command=self.disconnect_db).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Test Connection", command=self.test_connection).pack(side=tk.LEFT, padx=5)
    
    def setup_table_selection_tab(self):
        """Setup Table Selection Tab"""
        selection_tab = ttk.Frame(self.notebook)
        self.notebook.add(selection_tab, text="Table Selection")
        
        # Main frame with scrollbar
        main_frame = ttk.Frame(selection_tab)
        main_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Control buttons
        control_frame = ttk.Frame(main_frame)
        control_frame.pack(fill='x', pady=(0, 10))
        
        ttk.Button(control_frame, text="Select All", command=self.select_all_tables).pack(side=tk.LEFT, padx=2)
        ttk.Button(control_frame, text="Clear All", command=self.clear_all_tables).pack(side=tk.LEFT, padx=2)
        ttk.Button(control_frame, text="Refresh", command=self.refresh_tables).pack(side=tk.LEFT, padx=2)
        
        # Selected count label
        self.selected_count_label = ttk.Label(control_frame, text="0 tables selected", font=('Arial', 10))
        self.selected_count_label.pack(side=tk.RIGHT, padx=10)
        
        # Create scrollable canvas for tables
        canvas = tk.Canvas(main_frame, highlightthickness=0)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        self.checkbox_container = ttk.Frame(canvas)
        
        canvas.pack(side=tk.LEFT, fill='both', expand=True)
        scrollbar.pack(side=tk.RIGHT, fill='y')
        
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.create_window((0, 0), window=self.checkbox_container, anchor="nw")
        
        # Configure scrolling
        def on_frame_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
        
        self.checkbox_container.bind("<Configure>", on_frame_configure)
        
        # Safe mouse wheel binding
        def on_mousewheel(event):
            try:
                canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            except tk.TclError:
                pass
        
        canvas.bind("<MouseWheel>", on_mousewheel)
    
    def setup_position_mapping_tab(self):
        """Setup Position Mapping Tab"""
        mapping_tab = ttk.Frame(self.notebook)
        self.notebook.add(mapping_tab, text="Position Mapping")
        
        # Main frame
        main_frame = ttk.Frame(mapping_tab, padding="20")
        main_frame.pack(fill='both', expand=True)
        
        # Title
        ttk.Label(main_frame, text="Position Mapping Configuration", 
                 font=('Arial', 16, 'bold')).pack(pady=(0, 20))
        
        # Template section
        template_frame = ttk.LabelFrame(main_frame, text="Template", padding="10")
        template_frame.pack(fill='x', pady=(0, 20))
        
        ttk.Button(template_frame, text="Select Template", 
                  command=self.select_template).pack(side=tk.LEFT, padx=5)
        self.template_label = ttk.Label(template_frame, text="No template selected")
        self.template_label.pack(side=tk.LEFT, padx=10)
        ttk.Button(template_frame, text="Clear Template", 
                  command=self.clear_template).pack(side=tk.LEFT, padx=5)
        
        # Template info label
        self.template_info_label = ttk.Label(template_frame, text="", font=('Arial', 9))
        self.template_info_label.pack(side=tk.LEFT, padx=20)
        
        # Merge rules section
        merge_frame = ttk.LabelFrame(main_frame, text="Merge Cell Rules (Optional)", padding="10")
        merge_frame.pack(fill='x', pady=(0, 20))
        
        ttk.Label(merge_frame, text="Enter merge ranges (one per line): SheetName!StartCell:EndCell", 
                 font=('Arial', 9)).pack(anchor='w', pady=(0, 5))
        
        self.merge_rules_text = tk.Text(merge_frame, height=3, width=80, font=('Consolas', 9))
        self.merge_rules_text.pack(fill='x', pady=(0, 5))
        
        example_label = ttk.Label(merge_frame, 
                                 text="Example: Sheet1!B4:D4  (merges B4, C4, D4)\nExample: Sheet1!A1:C1  (merges A1, B1, C1)",
                                 font=('Arial', 8), foreground='blue')
        example_label.pack(anchor='w')
        
        # Instructions
        instr_frame = ttk.LabelFrame(main_frame, text="Instructions", padding="10")
        instr_frame.pack(fill='x', pady=(0, 20))
        
        instructions = [
            "1. First, select your Excel template above",
            "2. Optionally add merge cell rules (e.g., Sheet1!B4:D4)",
            "3. Configure position mappings for each table",
            "4. For each database column, specify:",
            "   • Which sheet (or 'All Sheets')",
            "   • Exact cell location (e.g., B4, C4, D4)",
            "   • Apply to: Choose 'All Sheets', 'This Sheet Only', or 'Select Sheets'",
            "5. For merged cells: Always write to top-left cell (e.g., B4 for B4:D4 merged)"
        ]
        
        for instr in instructions:
            ttk.Label(instr_frame, text=instr, font=('Arial', 9)).pack(anchor='w', pady=1)
        
        # Configuration button
        self.config_btn = ttk.Button(main_frame, text="Configure Position Mappings", 
                                    command=self.configure_positions, state='disabled')
        self.config_btn.pack(pady=10)
        
        # Current mappings display
        mapping_frame = ttk.LabelFrame(main_frame, text="Current Mappings", padding="10")
        mapping_frame.pack(fill='both', expand=True)
        
        self.mapping_text = scrolledtext.ScrolledText(mapping_frame, height=15, wrap=tk.WORD, font=('Consolas', 9))
        self.mapping_text.pack(fill='both', expand=True)
        self.update_mapping_display()
    
    def setup_export_tab(self):
        """Setup Export Tab"""
        export_tab = ttk.Frame(self.notebook)
        self.notebook.add(export_tab, text="Export")
        
        # Main frame
        main_frame = ttk.Frame(export_tab, padding="20")
        main_frame.pack(fill='both', expand=True)
        
        # Title
        ttk.Label(main_frame, text="Export Options", font=('Arial', 16, 'bold')).pack(pady=(0, 20))
        
        # Export options
        options_frame = ttk.LabelFrame(main_frame, text="Export Settings", padding="10")
        options_frame.pack(fill='x', pady=(0, 20))
        
        ttk.Label(options_frame, text="Row limit (0 = all):").pack(anchor='w', pady=2)
        self.row_limit_var = tk.StringVar(value="0")
        ttk.Entry(options_frame, textvariable=self.row_limit_var, width=10).pack(anchor='w', pady=2)
        
        # Export buttons
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(pady=20)
        
        self.export_btn = ttk.Button(btn_frame, text="Export to New Excel", 
                                    command=self.export_new_excel, state='normal')
        self.export_btn.pack(side=tk.LEFT, padx=5)
        
        self.template_export_btn = ttk.Button(btn_frame, text="Export to Template", 
                                             command=self.export_to_template, state='disabled')
        self.template_export_btn.pack(side=tk.LEFT, padx=5)
        
        # Log area
        log_frame = ttk.LabelFrame(main_frame, text="Export Log", padding="10")
        log_frame.pack(fill='both', expand=True)
        
        self.log_text = scrolledtext.ScrolledText(log_frame, height=10, font=('Consolas', 9))
        self.log_text.pack(fill='both', expand=True)
        
        # Configure log tags
        self.log_text.tag_configure('success', foreground='green', font=('Consolas', 9, 'bold'))
        self.log_text.tag_configure('error', foreground='red', font=('Consolas', 9, 'bold'))
        self.log_text.tag_configure('info', foreground='blue', font=('Consolas', 9))
        self.log_text.tag_configure('warning', foreground='orange', font=('Consolas', 9))
    
    def test_connection(self):
        """Test database connection"""
        def test():
            self.status_bar.config(text="Testing connection...")
            
            try:
                success, message = self.db.connect(
                    server=self.server_var.get(),
                    database=self.database_var.get(),
                    use_windows_auth=True
                )
                
                if success:
                    self.status_bar.config(text="Connection test successful")
                    self.db.disconnect()
                    messagebox.showinfo("Connection Test", "✅ Connection successful!")
                    self.log("Connection test successful", 'success')
                else:
                    self.status_bar.config(text="Connection test failed")
                    messagebox.showerror("Connection Test", f"❌ Connection failed:\n{message}")
                    self.log(f"Connection test failed: {message}", 'error')
                    
            except Exception as e:
                self.status_bar.config(text=f"Error: {str(e)}")
                messagebox.showerror("Connection Test", f"❌ Error during connection test:\n{str(e)}")
                self.log(f"Connection test error: {str(e)}", 'error')
        
        threading.Thread(target=test, daemon=True).start()
    
    def connect_db(self):
        """Connect to database"""
        def connect():
            self.status_bar.config(text="Connecting...")
            
            try:
                success, message = self.db.connect(
                    server=self.server_var.get(),
                    database=self.database_var.get(),
                    use_windows_auth=True
                )
                
                if success:
                    self.status_bar.config(text="Connected successfully")
                    self.refresh_tables()
                    self.log("✅ Database connected successfully", 'success')
                else:
                    self.status_bar.config(text=f"Connection failed: {message}")
                    messagebox.showerror("Connection Error", message)
                    self.log(f"❌ Connection failed: {message}", 'error')
                    
            except Exception as e:
                self.status_bar.config(text=f"Error: {str(e)}")
                messagebox.showerror("Connection Error", f"Error during connection:\n{str(e)}")
                self.log(f"❌ Connection error: {str(e)}", 'error')
        
        threading.Thread(target=connect, daemon=True).start()
    
    def disconnect_db(self):
        """Disconnect from database"""
        try:
            self.db.disconnect()
            self.status_bar.config(text="Disconnected")
            self.clear_table_checkboxes()
            self.selected_tables.clear()
            self.selected_count_label.config(text="0 tables selected")
            self.log("🔌 Disconnected from database", 'info')
        except Exception as e:
            self.log(f"❌ Error during disconnect: {str(e)}", 'error')
    
    def refresh_tables(self):
        """Refresh list of tables"""
        if not self.db.connected:
            messagebox.showwarning("Not Connected", "Please connect to database first")
            return
        
        def refresh():
            self.status_bar.config(text="Loading tables...")
            
            try:
                tables = self.db.get_tables()
                self.create_table_checkboxes(tables)
                self.status_bar.config(text=f"Loaded {len(tables)} tables")
                self.selected_count_label.config(text=f"{len(self.selected_tables)} tables selected")
                self.log(f"✅ Loaded {len(tables)} tables", 'success')
            except Exception as e:
                self.status_bar.config(text=f"Error loading tables: {str(e)}")
                self.log(f"❌ Error loading tables: {str(e)}", 'error')
        
        threading.Thread(target=refresh, daemon=True).start()
    
    def create_table_checkboxes(self, tables: List[str]):
        """Create checkboxes for table selection"""
        # Clear existing checkboxes
        self.clear_table_checkboxes()
        self.table_checkboxes.clear()
        
        # Create new checkboxes
        for i, table in enumerate(tables):
            var = tk.BooleanVar(value=False)
            self.table_checkboxes[table] = var
            
            cb_frame = ttk.Frame(self.checkbox_container)
            cb_frame.pack(fill='x', padx=5, pady=2)
            
            cb = ttk.Checkbutton(cb_frame, text=table, variable=var,
                                command=self.update_selected_count)
            cb.pack(anchor='w')
        
        self.update_selected_count()
    
    def clear_table_checkboxes(self):
        """Clear all table checkboxes"""
        for widget in self.checkbox_container.winfo_children():
            widget.destroy()
    
    def select_all_tables(self):
        """Select all tables"""
        for var in self.table_checkboxes.values():
            var.set(True)
        self.update_selected_count()
        self.log("✅ Selected all tables", 'info')
    
    def clear_all_tables(self):
        """Clear all table selections"""
        for var in self.table_checkboxes.values():
            var.set(False)
        self.update_selected_count()
        self.log("🗑️ Cleared all table selections", 'info')
    
    def update_selected_count(self):
        """Update selected tables count"""
        self.selected_tables = [table for table, var in self.table_checkboxes.items() if var.get()]
        count = len(self.selected_tables)
        self.selected_count_label.config(text=f"{count} table{'s' if count != 1 else ''} selected")
        
        # Enable/disable buttons based on selection
        if self.selected_tables and self.template_path:
            self.config_btn.config(state='normal')
            self.template_export_btn.config(state='normal')
        else:
            self.config_btn.config(state='disabled')
            if not self.template_path:
                self.template_export_btn.config(state='disabled')
    
    def select_template(self):
        """Select an Excel template file"""
        filetypes = [("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        filename = filedialog.askopenfilename(title="Select Excel Template", filetypes=filetypes)
        
        if filename:
            self.template_path = filename
            self.template_label.config(text=os.path.basename(filename))
            
            # Get sheet names from template
            try:
                wb = load_workbook(filename, read_only=True)
                self.template_sheets = wb.sheetnames
                sheet_info = f"{len(self.template_sheets)} sheets: {', '.join(self.template_sheets[:3])}"
                if len(self.template_sheets) > 3:
                    sheet_info += f" (+{len(self.template_sheets)-3} more)"
                self.template_info_label.config(text=sheet_info)
                
                self.log(f"✅ Template selected: {os.path.basename(filename)} with {len(self.template_sheets)} sheets", 'success')
            except Exception as e:
                self.log(f"❌ Error reading template: {str(e)}", 'error')
                self.template_sheets = []
                self.template_info_label.config(text="")
            
            # Enable buttons if tables are selected
            if self.selected_tables:
                self.config_btn.config(state='normal')
                self.template_export_btn.config(state='normal')
    
    def clear_template(self):
        """Clear the selected template"""
        self.template_path = None
        self.template_sheets = []
        self.template_label.config(text="No template selected")
        self.template_info_label.config(text="")
        self.config_btn.config(state='disabled')
        self.template_export_btn.config(state='disabled')
        self.log("🗑️ Template cleared", 'info')
    
    def configure_positions(self):
        """Configure position mappings for selected tables"""
        if not self.selected_tables:
            messagebox.showwarning("No Selection", "Please select tables first")
            return
        
        if not self.template_path:
            messagebox.showwarning("No Template", "Please select a template first")
            return
        
        if not self.template_sheets:
            messagebox.showerror("No Sheets", "Template has no sheets or could not be read")
            return
        
        # Get merge rules from text box
        try:
            merge_text = self.merge_rules_text.get("1.0", tk.END).strip()
            self.merge_rules = [line.strip() for line in merge_text.splitlines() if line.strip()]
            if self.merge_rules:
                self.log(f"Added {len(self.merge_rules)} merge rules", 'info')
        except:
            self.merge_rules = []
        
        # Configure each selected table
        for table_name in self.selected_tables:
            # Get database columns for this table
            db_columns = self.db.get_table_columns(table_name)
            
            if not db_columns:
                self.log(f"⚠️ No columns found for table: {table_name}", 'warning')
                continue
            
            # Show mapping dialog
            dialog = PositionMappingDialog(self.root, table_name, db_columns, self.template_sheets)
            self.root.wait_window(dialog.dialog)
            
            # Get mappings
            mappings = dialog.get_mappings()
            if mappings is None:
                continue  # User cancelled
            
            # Create or update table configuration
            if table_name not in self.table_configs:
                self.table_configs[table_name] = TableConfig(
                    table_name=table_name,
                    display_name=self.db.get_display_name(table_name),
                    start_row=0,
                    start_col="",
                    sheet_name=self.template_sheets[0] if self.template_sheets else "",
                    column_mappings={},
                    apply_to_all_sheets=False,
                    selected_sheets=[]
                )
            
            # Add column mappings
            for column_name, (sheet_name, cell_reference, apply_all, target_sheets) in mappings.items():
                self.table_configs[table_name].column_mappings[column_name] = CellMapping(
                    table_name=table_name,
                    column_name=column_name,
                    template_sheet=sheet_name,
                    template_cell=cell_reference,
                    apply_to_all_sheets=apply_all,
                    selected_sheets=target_sheets
                )
            
            self.log(f"✅ Configured {len(mappings)} column mappings for {table_name}", 'success')
        
        # Update display
        self.update_mapping_display()
        messagebox.showinfo("Configuration Complete", 
                          f"✅ Position mappings configured for {len(self.selected_tables)} tables")
    
    def update_mapping_display(self):
        """Update position mapping display"""
        self.mapping_text.delete(1.0, tk.END)
        
        if not self.table_configs:
            self.mapping_text.insert(1.0, "No position mappings configured yet.\n\n")
            self.mapping_text.insert(tk.END, "Click 'Configure Position Mappings' to set up mappings.")
        else:
            self.mapping_text.insert(tk.END, "✅ POSITION MAPPINGS CONFIGURED\n")
            self.mapping_text.insert(tk.END, "="*60 + "\n\n")
            
            # Show merge rules
            if self.merge_rules:
                self.mapping_text.insert(tk.END, "Merge Rules:\n")
                for rule in self.merge_rules:
                    self.mapping_text.insert(tk.END, f"  • {rule}\n")
                self.mapping_text.insert(tk.END, "\n")
            
            # Show table mappings
            for table_name, config in self.table_configs.items():
                self.mapping_text.insert(tk.END, f"📄 {config.display_name}\n")
                self.mapping_text.insert(tk.END, f"   Total Mappings: {len(config.column_mappings)}\n")
                
                if config.column_mappings:
                    self.mapping_text.insert(tk.END, "   Column Mappings:\n")
                    for col_name, cell_mapping in config.column_mappings.items():
                        if cell_mapping.apply_to_all_sheets:
                            apply_info = " (All Sheets)"
                        elif cell_mapping.selected_sheets:
                            sheets = cell_mapping.selected_sheets
                            if len(sheets) <= 3:
                                apply_info = f" (Sheets: {', '.join(sheets)})"
                            else:
                                apply_info = f" (Sheets: {', '.join(sheets[:3])} +{len(sheets)-3} more)"
                        else:
                            apply_info = f" (Sheet: {cell_mapping.template_sheet})"
                        
                        self.mapping_text.insert(tk.END, f"     • {col_name} → {cell_mapping.template_cell}{apply_info}\n")
                
                self.mapping_text.insert(tk.END, "\n")
    
    def fetch_table_data(self, limit: int = 0) -> Dict:
        """Fetch data for all selected tables"""
        tables_data = {}
        
        for table in self.selected_tables:
            try:
                data = self.db.fetch_table_data(table, limit if limit > 0 else None)
                tables_data[table] = data
                
                if data['success']:
                    self.log(f"✅ {data['display_name']}: {data['row_count']} rows", 'success')
                else:
                    self.log(f"❌ {data['display_name']}: {data.get('error', 'Unknown error')}", 'error')
                    
            except Exception as e:
                tables_data[table] = {'success': False, 'error': str(e)}
                self.log(f"❌ {table}: {str(e)}", 'error')
        
        return tables_data
    
    def export_new_excel(self):
        """Export selected tables to new Excel file"""
        if not self.selected_tables:
            messagebox.showwarning("No Selection", "Please select tables first")
            return
        
        # Ask for save location
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"TableExport_{timestamp}.xlsx"
        
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=filename,
            title="Save Excel File"
        )
        
        if not file_path:
            return
        
        def export():
            self.status_bar.config(text="Exporting...")
            self.log("Starting export...", 'info')
            
            try:
                # Get row limit
                try:
                    row_limit = int(self.row_limit_var.get())
                except:
                    row_limit = 0
                
                # Fetch data
                tables_data = self.fetch_table_data(row_limit)
                
                # Export to Excel
                self.exporter.export_tables_to_excel(tables_data, file_path)
                
                self.status_bar.config(text="Export completed")
                self.log(f"✅ Export completed: {file_path}", 'success')
                
                # Show success message
                self.root.after(0, lambda: self.show_export_success(file_path))
                
            except Exception as e:
                error_msg = str(e)
                self.status_bar.config(text=f"Export failed: {error_msg}")
                self.log(f"❌ Export failed: {error_msg}", 'error')
                self.root.after(0, lambda: messagebox.showerror("Export Error", f"Failed to export:\n{error_msg}"))
        
        threading.Thread(target=export, daemon=True).start()
    
    def export_to_template(self):
        """Export selected tables to template using position mappings"""
        if not self.selected_tables:
            messagebox.showwarning("No Selection", "Please select tables first")
            return
        
        if not self.template_path:
            messagebox.showwarning("No Template", "Please select a template first")
            return
        
        if not self.table_configs:
            if not messagebox.askyesno("No Mappings", 
                                      "No position mappings configured. Export anyway?"):
                return
        
        # Get merge rules
        try:
            merge_text = self.merge_rules_text.get("1.0", tk.END).strip()
            merge_rules = [line.strip() for line in merge_text.splitlines() if line.strip()]
        except:
            merge_rules = []
        
        # Ask for save location
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"TemplateExport_{timestamp}.xlsx"
        
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=filename,
            title="Save Excel File"
        )
        
        if not file_path:
            return
        
        def export():
            self.status_bar.config(text="Exporting to template...")
            self.log("Starting template export...", 'info')
            
            try:
                # Get row limit
                try:
                    row_limit = int(self.row_limit_var.get())
                except:
                    row_limit = 0
                
                # Fetch data
                tables_data = self.fetch_table_data(row_limit)
                
                # Export to template with merge rules
                self.exporter.export_tables_to_template(
                    tables_data=tables_data,
                    template_path=self.template_path,
                    table_configs=self.table_configs,
                    output_path=file_path,
                    merge_rules=merge_rules
                )
                
                self.status_bar.config(text="Export completed")
                self.log(f"✅ Template export completed: {file_path}", 'success')
                
                # Show success message
                self.root.after(0, lambda: self.show_export_success(file_path))
                
            except Exception as e:
                error_msg = str(e)
                self.status_bar.config(text=f"Export failed: {error_msg}")
                self.log(f"❌ Template export failed: {error_msg}", 'error')
                self.root.after(0, lambda: messagebox.showerror("Export Error", 
                    f"Failed to export template:\n{error_msg}\n\n"
                    "Possible causes:\n"
                    "1. Template file is open in Excel - close it and try again\n"
                    "2. Invalid cell references in mappings\n"
                    "3. File permissions issue"))
        
        threading.Thread(target=export, daemon=True).start()
    
    def show_export_success(self, file_path: str):
        """Show export success dialog"""
        file_name = os.path.basename(file_path)
        
        dialog = tk.Toplevel(self.root)
        dialog.title("Export Successful")
        dialog.geometry("400x250")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Center dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() - dialog.winfo_width()) // 2
        y = (dialog.winfo_screenheight() - dialog.winfo_height()) // 2
        dialog.geometry(f"+{x}+{y}")
        
        # Success message
        ttk.Label(dialog, text="✅", font=('Arial', 32), foreground='green').pack(pady=(20, 10))
        ttk.Label(dialog, text="Excel File Created Successfully!", font=('Arial', 12, 'bold')).pack()
        
        # File info
        file_size = os.path.getsize(file_path) / 1024  # KB
        ttk.Label(dialog, text=f"File: {file_name}").pack(pady=5)
        ttk.Label(dialog, text=f"Size: {file_size:.1f} KB").pack(pady=5)
        
        # Buttons
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=20)
        
        ttk.Button(btn_frame, text="Open File", 
                  command=lambda: [os.startfile(file_path), dialog.destroy()],
                  width=12).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Open Folder", 
                  command=lambda: [os.startfile(os.path.dirname(file_path)), dialog.destroy()],
                  width=12).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="OK", 
                  command=dialog.destroy, width=8).pack(side=tk.RIGHT, padx=5)
    
    def log(self, message: str, tag: str = 'info'):
        """Add message to log"""
        timestamp = datetime.now().strftime('%H:%M:%S')
        self.log_text.insert(tk.END, f"[{timestamp}] {message}\n", tag)
        self.log_text.see(tk.END)
    
    def on_closing(self):
        """Handle window closing"""
        if messagebox.askokcancel("Quit", "Do you want to quit the application?"):
            if self.db.connected:
                self.db.disconnect()
            self.root.destroy()

# ============================================================================
# MAIN ENTRY POINT
# ============================================================================

def main():
    """Main function to run the application"""
    root = tk.Tk()
    
    # Create application
    app = MultiTableExporterApp(root)
    
    # Center window on screen
    root.update_idletasks()
    width = root.winfo_width()
    height = root.winfo_height()
    x = (root.winfo_screenwidth() // 2) - (width // 2)
    y = (root.winfo_screenheight() // 2) - (height // 2)
    root.geometry(f'{width}x{height}+{x}+{y}')
    
    # Start main loop
    root.mainloop()

if __name__ == "__main__":
    main()