"""
Excel Table Exporter - Complete Streamlit Application
Version: 3.2.0 - Fixed Template Export Issue & No Column Names in Data Fetch
Description: Export SQL Server tables to Excel templates with position mapping
With advanced filtering and multi-sheet support
"""

import streamlit as st
import pandas as pd
import os
import sys
import logging
from datetime import datetime, timedelta
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
import traceback
import shutil
import re
import tempfile
import base64
from io import BytesIO
from typing import Dict, List, Optional, Any, Tuple
import json

# Database imports - using SQLAlchemy for compatibility
try:
    from sqlalchemy import create_engine, text
    from sqlalchemy.exc import SQLAlchemyError
    SQLALCHEMY_AVAILABLE = True
except ImportError:
    SQLALCHEMY_AVAILABLE = False
    st.error("SQLAlchemy not available. Please install it.")

# ============================================================================
# LOGGING SETUP
# ============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('excel_exporter.log', encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# ============================================================================
# DATABASE MANAGER (Based on Tkinter XYZ.py logic)
# ============================================================================

class DatabaseManager:
    """Manages database connections using SQLAlchemy with Windows Authentication"""
    
    def __init__(self):
        self.engine = None
        self.connected = False
        self.server = None
        self.database = None
        self.driver = None
    
    def connect(self, server: str, database: str, driver: str = "auto") -> Tuple[bool, str]:
        """Connect to SQL Server with Windows Authentication only"""
        if not SQLALCHEMY_AVAILABLE:
            return False, "SQLAlchemy not available"
        
        try:
            logger.info(f"Connecting to {server}.{database} with Windows Authentication")
            
            # Try different connection methods (from XYZ.py logic)
            connection_methods = []
            
            # Method 1: pyodbc with ODBC Driver 17
            connection_methods.append({
                'name': 'pyodbc',
                'string': f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={server};DATABASE={database};Trusted_Connection=yes;TrustServerCertificate=yes;",
                'engine_string': "mssql+pyodbc:///?odbc_connect={}"
            })
            
            # Method 2: pymssql
            connection_methods.append({
                'name': 'pymssql',
                'string': f"mssql+pymssql://{server}/{database}",
                'engine_string': None
            })
            
            # Try each method
            for method in connection_methods:
                try:
                    logger.info(f"Trying connection method: {method['name']}")
                    
                    if method['name'] == 'pyodbc':
                        connection_string = method['string']
                        engine_string = method['engine_string'].format(connection_string)
                        self.engine = create_engine(
                            engine_string,
                            pool_pre_ping=True,
                            echo=False
                        )
                    else:
                        self.engine = create_engine(
                            method['string'],
                            pool_pre_ping=True,
                            echo=False
                        )
                    
                    # Test connection
                    with self.engine.connect() as conn:
                        conn.execute(text("SELECT 1"))
                    
                    self.connected = True
                    self.server = server
                    self.database = database
                    self.driver = method['name']
                    
                    logger.info(f"✅ Connection successful using {method['name']}")
                    return True, f"Connection successful using {method['name']}"
                    
                except Exception as e:
                    logger.warning(f"Connection method {method['name']} failed: {str(e)}")
                    continue
            
            return False, "All connection attempts failed"
            
        except Exception as e:
            error_msg = f"Connection error: {str(e)}"
            logger.error(error_msg)
            return False, error_msg
    
    def disconnect(self):
        """Disconnect from database"""
        try:
            if self.engine:
                self.engine.dispose()
            logger.info("Disconnected")
        except Exception as e:
            logger.error(f"Error disconnecting: {e}")
        finally:
            self.engine = None
            self.connected = False
            self.server = None
            self.database = None
            self.driver = None
    
    def get_tables(self) -> List[str]:
        """Get list of tables"""
        try:
            query = """
            SELECT TABLE_NAME 
            FROM INFORMATION_SCHEMA.TABLES 
            WHERE TABLE_TYPE = 'BASE TABLE'
            ORDER BY TABLE_NAME
            """
            
            with self.engine.connect() as conn:
                result = conn.execute(text(query))
                tables = [row[0] for row in result.fetchall()]
            
            logger.info(f"Found {len(tables)} tables")
            return tables
            
        except Exception as e:
            logger.error(f"Error getting tables: {e}")
            return []
    
    def get_table_columns(self, table_name: str) -> List[str]:
        """Get column names for a table"""
        try:
            query = """
            SELECT COLUMN_NAME 
            FROM INFORMATION_SCHEMA.COLUMNS 
            WHERE TABLE_NAME = :table_name
            ORDER BY ORDINAL_POSITION
            """
            
            with self.engine.connect() as conn:
                result = conn.execute(text(query), {"table_name": table_name})
                columns = [row[0] for row in result.fetchall()]
            
            logger.info(f"Found {len(columns)} columns in {table_name}")
            return columns
            
        except Exception as e:
            logger.error(f"Error getting columns: {e}")
            return []
    
    def get_batches_from_table(self, table_name: str) -> List[str]:
        """Get distinct batch names from a table (from XYZ.py)"""
        try:
            columns = self.get_table_columns(table_name)
            batch_column = None
            
            # Look for common batch column names
            batch_keywords = ['BATCH', 'BATCH_NAME', 'BATCH_NUMBER', 'BATCH_NO', 'BATCHID']
            for col in columns:
                if any(keyword in col.upper() for keyword in batch_keywords):
                    batch_column = col
                    break
            
            if not batch_column:
                logger.warning(f"No batch column found in table {table_name}")
                return []
            
            # Get distinct batches
            query = f"SELECT DISTINCT [{batch_column}] FROM [{table_name}] WHERE [{batch_column}] IS NOT NULL ORDER BY [{batch_column}]"
            
            with self.engine.connect() as conn:
                result = conn.execute(text(query))
                batches = [str(row[0]) for row in result.fetchall()]
            
            logger.info(f"Retrieved {len(batches)} batches from {table_name}")
            return batches
            
        except Exception as e:
            logger.error(f"Error getting batches from {table_name}: {e}")
            return []
    
    def get_time_columns(self, table_name: str) -> List[str]:
        """Get time-related columns from a table (from XYZ.py)"""
        try:
            columns = self.get_table_columns(table_name)
            time_columns = []
            
            # Look for time-related columns
            time_keywords = ['TIME', 'DATE', 'TIMESTAMP', 'DATETIME', 'START', 'STOP', 'END']
            for col in columns:
                if any(keyword in col.upper() for keyword in time_keywords):
                    time_columns.append(col)
            
            logger.info(f"Found {len(time_columns)} time columns in {table_name}")
            return time_columns
            
        except Exception as e:
            logger.error(f"Error getting time columns from {table_name}: {e}")
            return []
    
    def fetch_filtered_data(self, table_name: str, batch_name: str = None, 
                          start_time: datetime = None, end_time: datetime = None,
                          limit: int = None) -> Dict:
        """Fetch data from a table with filters for batch and time range - VALUES ONLY (NO COLUMN NAMES)"""
        try:
            logger.info(f"[FETCH] Fetching filtered data from table: {table_name}")
            
            display_name = self.get_display_name(table_name)
            temp_columns = self.get_table_columns(table_name)
            
            # Build WHERE clause for filters
            where_clauses = []
            params = {}
            
            # Add batch filter if specified
            if batch_name:
                # Find batch column
                batch_column = None
                batch_keywords = ['BATCH', 'BATCH_NAME', 'BATCH_NUMBER', 'BATCH_NO', 'BATCHID']
                for col in temp_columns:
                    if any(keyword in col.upper() for keyword in batch_keywords):
                        batch_column = col
                        break
                
                if batch_column:
                    where_clauses.append(f"[{batch_column}] = :batch_name")
                    params["batch_name"] = batch_name
                    logger.info(f"Filtering by batch: {batch_name} in column {batch_column}")
                else:
                    logger.warning(f"No batch column found for filtering")
            
            # Add time range filter if specified
            if start_time or end_time:
                # Find time column
                time_column = None
                time_keywords = ['TIME', 'TIMESTAMP', 'DATETIME', 'DATE_TIME', 'CREATED_AT']
                for col in temp_columns:
                    if any(keyword in col.upper() for keyword in time_keywords):
                        time_column = col
                        break
                
                if time_column:
                    if start_time:
                        where_clauses.append(f"[{time_column}] >= :start_time")
                        params["start_time"] = start_time
                    if end_time:
                        where_clauses.append(f"[{time_column}] <= :end_time")
                        params["end_time"] = end_time
                    logger.info(f"Filtering by time range in column {time_column}")
                else:
                    logger.warning(f"No time column found for filtering")
            
            # Build query - SELECT * to get all data
            if limit and limit > 0:
                query = f"SELECT TOP ({limit}) * FROM [{table_name}]"
            else:
                query = f"SELECT * FROM [{table_name}]"
            
            if where_clauses:
                query += " WHERE " + " AND ".join(where_clauses)
            
            # Add order by time if available
            time_columns = self.get_time_columns(table_name)
            if time_columns:
                query += f" ORDER BY [{time_columns[0]}]"
            
            logger.debug(f"Executing query: {query}")
            logger.debug(f"Parameters: {params}")
            
            # Execute query
            with self.engine.connect() as conn:
                df = pd.read_sql_query(text(query), conn, params=params)
            
            # Convert to list of lists - PURE VALUES ONLY (NO COLUMN NAMES)
            data = []
            for _, row in df.iterrows():
                row_list = []
                for value in row:
                    if pd.isna(value):
                        row_list.append("")
                    elif isinstance(value, datetime):
                        row_list.append(value.strftime('%Y-%m-%d %H:%M:%S'))
                    elif isinstance(value, timedelta):
                        total_seconds = int(value.total_seconds())
                        hours, remainder = divmod(total_seconds, 3600)
                        minutes, seconds = divmod(remainder, 60)
                        row_list.append(f"{hours:02d}:{minutes:02d}:{seconds:02d}")
                    else:
                        row_list.append(str(value))
                data.append(row_list)
            
            row_count = len(data)
            
            if data:
                logger.debug(f"Sample filtered row from {table_name}: First 3 values - {data[0][:3]}")
            
            logger.info(f"[OK] Fetched {row_count} filtered rows from {table_name} (VALUES ONLY, NO COLUMN NAMES)")
            
            # Return ONLY data values, no column names
            return {
                'success': True,
                'display_name': display_name,
                'table_name': table_name,
                'data': data,  # Only values, no column names
                'row_count': row_count,
                'filters_applied': {
                    'batch': batch_name,
                    'start_time': start_time,
                    'end_time': end_time
                }
            }
            
        except Exception as e:
            error_msg = f"Error fetching filtered data from {table_name}: {str(e)}"
            logger.error(f"[ERROR] {error_msg}")
            logger.error(traceback.format_exc())
            return {
                'success': False,
                'error': error_msg,
                'display_name': self.get_display_name(table_name),
                'table_name': table_name,
                'data': [],
                'row_count': 0
            }
    
    def get_display_name(self, table_name: str) -> str:
        """Convert table name to display name"""
        display_name = table_name.replace('_', ' ').title()
        
        replacements = {
            'Tbl': '',
            'Table': '',
            'Vw': 'View: ',
            'Vw_': 'View: ',
            'View': 'View: '
        }
        
        for old, new in replacements.items():
            if display_name.startswith(old):
                display_name = display_name.replace(old, new, 1).strip()
        
        return display_name if display_name else table_name

# ============================================================================
# DATA CLASSES (from XYZ.py)
# ============================================================================

from dataclasses import dataclass, field
from typing import Dict, List

@dataclass
class CellMapping:
    """Mapping information for a single cell"""
    table_name: str
    column_name: str
    template_sheet: str
    template_cell: str  # e.g., "A1", "B3", "C5"
    apply_to_all_sheets: bool = False
    selected_sheets: List[str] = field(default_factory=list)
    
@dataclass
class TableConfig:
    """Configuration for a data table"""
    table_name: str
    display_name: str
    start_row: int
    start_col: str
    sheet_name: str
    column_mappings: Dict[str, CellMapping] = field(default_factory=dict)
    apply_to_all_sheets: bool = False
    selected_sheets: List[str] = field(default_factory=list)

# ============================================================================
# EXCEL EXPORTER WITH SHEET SELECTION SUPPORT (FIXED VERSION)
# ============================================================================

class ExcelTableExporter:
    """Handles exporting tables to Excel with position mapping and merged cell support"""
    
    @staticmethod
    def export_tables_to_new_excel(tables_data: Dict, output_path: str) -> bool:
        """Export multiple tables to new Excel file (no column names in data)"""
        try:
            wb = Workbook()
            
            # Remove default sheet
            if wb.sheetnames:
                wb.remove(wb.active)
            
            # Create a sheet for each table
            for table_name, table_data in tables_data.items():
                if table_data.get('success', False):
                    # Create sheet with valid name
                    sheet_name = ExcelTableExporter.get_valid_sheet_name(table_data['display_name'])
                    ws = wb.create_sheet(title=sheet_name)
                    
                    # Write data directly (no headers since we don't have column names)
                    data = table_data['data']
                    for row_idx, row_data in enumerate(data, 1):  # Start from row 1
                        for col_idx, value in enumerate(row_data, 1):
                            ws.cell(row=row_idx, column=col_idx, value=value)
            
            wb.save(output_path)
            logger.info(f"✅ Excel file created: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error creating Excel: {str(e)}")
            return False
    
    @staticmethod
    def export_tables_to_template(tables_data: Dict, template_path: str, 
                                table_configs: Dict[str, TableConfig],
                                output_path: str,
                                merge_rules: List[str] = None) -> bool:
        """
        Export data into an existing template using position mappings.
        Template structure is kept AS IS, only values are filled in.
        """
        try:
            logger.info("="*60)
            logger.info("[START] STARTING TEMPLATE EXPORT")
            logger.info("="*60)
            logger.info(f"Template: {template_path}")
            logger.info(f"Output: {output_path}")
            logger.info(f"Tables to export: {list(tables_data.keys())}")
            
            # Check if we have data
            total_rows = sum(t.get('row_count', 0) for t in tables_data.values() if t.get('success', False))
            logger.info(f"Total rows to export: {total_rows}")
            
            if total_rows == 0:
                logger.warning("[WARNING] No data found to export!")
                return False
            
            # Make a copy of the template
            logger.info("[COPY] Copying template...")
            shutil.copy2(template_path, output_path)
            
            # Load the copied template
            logger.info("[LOAD] Loading template workbook...")
            wb = load_workbook(output_path)
            logger.info(f"[SHEETS] Workbook sheets: {wb.sheetnames}")
            
            # Apply user merge rules first (optional)
            if merge_rules:
                logger.info(f"Applying {len(merge_rules)} merge rules")
                for rule in merge_rules:
                    try:
                        if "!" in rule:
                            sheet_name, cell_range = rule.split("!", 1)
                            sheet_name = sheet_name.strip()
                            cell_range = cell_range.strip()
                            if sheet_name in wb.sheetnames:
                                wb[sheet_name].merge_cells(cell_range)
                                logger.info(f"[OK] Merged {sheet_name}!{cell_range}")
                    except Exception as e:
                        logger.warning(f"Failed to apply merge rule '{rule}': {e}")
            
            # Process each table
            for table_name, table_data in tables_data.items():
                logger.info("-"*40)
                logger.info(f"[PROCESS] Processing table: {table_name}")
                
                if not table_data.get('success', False):
                    logger.warning(f"[ERROR] Table {table_name} has error: {table_data.get('error')}")
                    continue
                
                if table_name not in table_configs:
                    logger.warning(f"[WARNING] No configuration found for table: {table_name}")
                    continue
                
                table_config = table_configs[table_name]
                logger.info(f"Table type: {'TABULAR' if table_config.start_row > 0 else 'HEADER'}")
                logger.info(f"Data: {table_data.get('row_count', 0)} rows (VALUES ONLY)")
                
                # Check if this is a header table with column mappings
                if table_config.column_mappings:
                    logger.info(f"Processing {len(table_config.column_mappings)} column mappings for header table")
                    
                    if table_data['data'] and len(table_data['data']) > 0:
                        first_row = table_data['data'][0]  # First data row
                        
                        for column_name, cell_mapping in table_config.column_mappings.items():
                            logger.debug(f"Mapping column: {column_name}")
                            
                            # Find which position this column is in the data
                            # For header tables, we need to match column names
                            column_index = -1
                            # Since we don't have column names in data, we use the order from column_mappings
                            try:
                                col_list = list(table_config.column_mappings.keys())
                                column_index = col_list.index(column_name)
                            except:
                                column_index = -1
                            
                            # Get value from first row
                            value = ""
                            if column_index >= 0 and column_index < len(first_row):
                                value = first_row[column_index]
                            
                            logger.debug(f"Value for {column_name}: {value}")
                            
                            # Determine which sheets to write to
                            sheets_to_write = []
                            if cell_mapping.apply_to_all_sheets or table_config.apply_to_all_sheets:
                                # Write to all sheets
                                sheets_to_write = wb.sheetnames
                            elif cell_mapping.selected_sheets:
                                # Write to selected sheets
                                sheets_to_write = [s for s in cell_mapping.selected_sheets if s in wb.sheetnames]
                            elif table_config.selected_sheets:
                                # Write to table's selected sheets
                                sheets_to_write = [s for s in table_config.selected_sheets if s in wb.sheetnames]
                            else:
                                # Write to specific sheet only
                                if cell_mapping.template_sheet in wb.sheetnames:
                                    sheets_to_write = [cell_mapping.template_sheet]
                            
                            # Write to each sheet
                            for sheet_name in sheets_to_write:
                                success = ExcelTableExporter.write_to_cell_safe(
                                    wb, 
                                    sheet_name, 
                                    cell_mapping.template_cell, 
                                    value
                                )
                                
                                if success:
                                    logger.debug(f"[OK] Wrote '{value}' to {sheet_name}!{cell_mapping.template_cell}")
                                else:
                                    logger.warning(f"[ERROR] Could not write to {sheet_name}!{cell_mapping.template_cell}")
                    else:
                        logger.warning(f"No data found for header table {table_name}")
                            
                # Write tabular data if start position is configured
                if table_config.start_row > 0 and table_config.start_col:
                    logger.info(f"Writing tabular data starting at {table_config.start_col}{table_config.start_row}")
                    
                    if not table_data['data']:
                        logger.warning(f"[WARNING] No data found for table {table_name}")
                        continue
                    
                    # Determine which sheets to write to
                    sheets_to_write = []
                    if table_config.apply_to_all_sheets:
                        sheets_to_write = wb.sheetnames
                    elif table_config.selected_sheets:
                        sheets_to_write = [s for s in table_config.selected_sheets if s in wb.sheetnames]
                    else:
                        if table_config.sheet_name in wb.sheetnames:
                            sheets_to_write = [table_config.sheet_name]
                    
                    logger.info(f"Writing to {len(sheets_to_write)} sheets")
                    
                    for sheet_name in sheets_to_write:
                        ws = wb[sheet_name]
                        start_col_idx = column_index_from_string(table_config.start_col)
                        
                        # Find first safe row
                        safe_row = ExcelTableExporter.find_safe_row_for_table(ws, table_config.start_row)
                        logger.info(f"Writing to sheet '{sheet_name}' starting at row {safe_row}, column {table_config.start_col}")
                        
                        # Write data (PURE VALUES ONLY - NO HEADERS)
                        data_rows = table_data['data']
                        logger.info(f"Writing {len(data_rows)} data rows")
                        
                        # Start data from the specified row directly (no headers since we don't have column names)
                        start_data_row = safe_row
                        
                        # Write data rows
                        for row_idx, row_data in enumerate(data_rows, start=0):
                            for col_idx, value in enumerate(row_data, start=0):
                                cell_col = start_col_idx + col_idx
                                cell_row = start_data_row + row_idx
                                cell_ref = f"{get_column_letter(cell_col)}{cell_row}"
                                
                                ExcelTableExporter.write_to_cell_safe(
                                    wb, sheet_name, cell_ref, value
                                )
                            
                            # Log progress every 10 rows
                            if (row_idx + 1) % 10 == 0:
                                logger.debug(f"  Processed {row_idx + 1} rows...")
            
            # Save workbook
            logger.info("[SAVE] Saving workbook...")
            wb.save(output_path)
            logger.info("="*60)
            logger.info("[OK] TEMPLATE EXPORT COMPLETED SUCCESSFULLY")
            logger.info("="*60)
            return True
            
        except Exception as e:
            logger.error("="*60)
            logger.error("[ERROR] TEMPLATE EXPORT FAILED")
            logger.error("="*60)
            logger.error(f"Error: {str(e)}")
            logger.error(traceback.format_exc())
            return False
    
    @staticmethod
    def write_to_cell_safe(wb, sheet_name: str, cell_ref: str, value: Any) -> bool:
        """Safely write to a cell, handling merged cells."""
        try:
            if sheet_name not in wb.sheetnames:
                logger.warning(f"Sheet '{sheet_name}' not found in workbook")
                return False
            
            ws = wb[sheet_name]
            
            # Parse cell reference
            col_letter = ''.join([c for c in cell_ref if c.isalpha()])
            row_num = int(''.join([c for c in cell_ref if c.isdigit()]))
            
            # Validate cell reference
            if not col_letter or not row_num:
                logger.warning(f"Invalid cell reference: {cell_ref}")
                return False
            
            col_num = column_index_from_string(col_letter)
            
            # Check if cell is part of a merged range
            for merged_range in ws.merged_cells.ranges:
                if (merged_range.min_row <= row_num <= merged_range.max_row and
                    merged_range.min_col <= col_num <= merged_range.max_col):
                    # Cell is in a merged range
                    # Try to write to the top-left cell of the merged range
                    top_left_cell = f"{get_column_letter(merged_range.min_col)}{merged_range.min_row}"
                    try:
                        ws[top_left_cell] = value
                        # Center alignment for merged cells
                        ws[top_left_cell].alignment = Alignment(horizontal='center', vertical='center')
                        logger.debug(f"[WRITE] Wrote to merged cell {top_left_cell} (original: {cell_ref})")
                        return True
                    except Exception as e:
                        logger.warning(f"Failed to write to merged cell {top_left_cell}: {e}")
                        return False
            
            # Cell is not merged, write normally
            ws[cell_ref] = value
            logger.debug(f"[WRITE] Wrote to cell {cell_ref}")
            return True
            
        except Exception as e:
            logger.error(f"[ERROR] Error writing to cell {sheet_name}!{cell_ref}: {e}")
            return False
    
    @staticmethod
    def find_safe_row_for_table(ws, start_row: int) -> int:
        """Find a safe row to start table data (not in merged cells)"""
        current_row = start_row
        
        # Check if start row is safe (not in merged cells in first 5 columns)
        for col in range(1, 6):  # Check first 5 columns
            cell_ref = f"{get_column_letter(col)}{current_row}"
            for merged_range in ws.merged_cells.ranges:
                if cell_ref in merged_range:
                    # Row is in merged range, try next row
                    current_row += 1
                    logger.debug(f"Row {current_row-1} is in merged range, trying row {current_row}")
                    return ExcelTableExporter.find_safe_row_for_table(ws, current_row)
        
        return current_row
    
    @staticmethod
    def get_valid_sheet_name(name: str) -> str:
        """Get valid Excel sheet name"""
        invalid_chars = ['\\', '/', '*', '?', ':', '[', ']']
        for char in invalid_chars:
            name = name.replace(char, '_')
        
        if len(name) > 31:
            name = name[:28] + "..."
        
        if not name.strip():
            name = "Sheet"
        
        return name[:31]

# ============================================================================
# STREAMLIT APPLICATION
# ============================================================================

def main():
    # Page configuration
    st.set_page_config(
        page_title="Excel Table Exporter Pro",
        page_icon="📊",
        layout="wide",
        initial_sidebar_state="expanded"
    )
    
    # Custom CSS
    st.markdown("""
    <style>
    .main-header {
        font-size: 2.5rem;
        color: #1E3A8A;
        margin-bottom: 1rem;
    }
    .step-box {
        background-color: #f8f9fa;
        padding: 1.5rem;
        border-radius: 10px;
        border-left: 5px solid #3B82F6;
        margin-bottom: 1.5rem;
    }
    .success-box {
        background-color: #d4edda;
        padding: 1rem;
        border-radius: 5px;
        border-left: 4px solid #28a745;
    }
    .warning-box {
        background-color: #fff3cd;
        padding: 1rem;
        border-radius: 5px;
        border-left: 4px solid #ffc107;
    }
    .info-box {
        background-color: #d1ecf1;
        padding: 1rem;
        border-radius: 5px;
        border-left: 4px solid #17a2b8;
    }
    .stDownloadButton {
        display: inline-block;
        padding: 0.5rem 1rem;
        background-color: #4CAF50;
        color: white;
        text-align: center;
        text-decoration: none;
        border-radius: 4px;
        font-weight: bold;
        margin: 10px 0;
    }
    .debug-info {
        background-color: #f8f9fa;
        padding: 10px;
        border-radius: 5px;
        margin: 10px 0;
        font-family: monospace;
        font-size: 0.9rem;
    }
    .table-config {
        background-color: #f0f7ff;
        padding: 1rem;
        border-radius: 5px;
        margin: 10px 0;
        border-left: 4px solid #4A90E2;
    }
    .merge-rules {
        background-color: #fff8e1;
        padding: 1rem;
        border-radius: 5px;
        margin: 10px 0;
        border-left: 4px solid #FFA000;
    }
    .filter-info {
        background-color: #e8f5e9;
        padding: 1rem;
        border-radius: 5px;
        margin: 10px 0;
        border-left: 4px solid #43A047;
    }
    .export-log {
        background-color: #f5f5f5;
        padding: 1rem;
        border-radius: 5px;
        margin: 10px 0;
        font-family: monospace;
        font-size: 0.8rem;
        max-height: 300px;
        overflow-y: auto;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # Title
    st.markdown('<h1 class="main-header">📊 Excel Table Exporter Pro</h1>', unsafe_allow_html=True)
    st.markdown("Advanced SQL Server to Excel exporter with filtering and multi-sheet support")
    
    # Initialize session state
    if 'db' not in st.session_state:
        st.session_state.db = DatabaseManager()
    if 'step' not in st.session_state:
        st.session_state.step = 1
    if 'tables_list' not in st.session_state:
        st.session_state.tables_list = []
    if 'selected_tables' not in st.session_state:
        st.session_state.selected_tables = []
    if 'template_path' not in st.session_state:
        st.session_state.template_path = None
    if 'template_sheets' not in st.session_state:
        st.session_state.template_sheets = []
    if 'table_configs' not in st.session_state:
        st.session_state.table_configs = {}
    if 'filters' not in st.session_state:
        st.session_state.filters = {}
    if 'merge_rules' not in st.session_state:
        st.session_state.merge_rules = []
    if 'export_log' not in st.session_state:
        st.session_state.export_log = []
    if 'row_limit' not in st.session_state:
        st.session_state.row_limit = 1000
    if 'configuring_positions' not in st.session_state:
        st.session_state.configuring_positions = False
    
    # Sidebar
    with st.sidebar:
        st.image("https://img.icons8.com/color/96/000000/microsoft-excel-2019.png", width=80)
        
        st.markdown("### Navigation")
        
        # Progress steps
        steps = [
            ("🔗", "Connection", 1),
            ("📋", "Table Selection", 2),
            ("📍", "Position Mapping", 3),
            ("⚙️", "Filters", 4),
            ("📤", "Export", 5)
        ]
        
        for icon, name, step_num in steps:
            if step_num == st.session_state.step:
                st.markdown(f"**{icon} {name}**")
            else:
                if st.button(f"{icon} {name}", key=f"nav_{step_num}", use_container_width=True):
                    st.session_state.step = step_num
                    st.rerun()
        
        st.divider()
        
        # Connection status
        if st.session_state.db.connected:
            st.success("✅ Connected")
            st.info(f"Database: {st.session_state.db.database}")
            if st.button("🔌 Disconnect", use_container_width=True):
                st.session_state.db.disconnect()
                st.session_state.tables_list = []
                st.session_state.selected_tables = []
                st.session_state.filters = {}
                st.session_state.step = 1
                st.rerun()
        else:
            st.warning("⚠️ Not Connected")
        
        # Selected tables
        if st.session_state.selected_tables:
            st.info(f"📋 {len(st.session_state.selected_tables)} tables selected")
        
        # Template status
        if st.session_state.template_path:
            st.info(f"📄 Template loaded")
        
        # Configurations
        if st.session_state.table_configs:
            st.info(f"📍 {len(st.session_state.table_configs)} mappings")
        
        # Filters
        if st.session_state.filters:
            st.info(f"⚙️ {len(st.session_state.filters)} filtered tables")
    
    # Main content based on current step
    if st.session_state.step == 1:
        show_connection_tab()
    elif st.session_state.step == 2:
        show_table_selection_tab()
    elif st.session_state.step == 3:
        show_position_mapping_tab()
    elif st.session_state.step == 4:
        show_filters_tab()
    elif st.session_state.step == 5:
        show_export_tab()

def show_connection_tab():
    """Connection tab"""
    st.markdown("## Step 1: Database Connection")
    
    with st.container():
        st.markdown('<div class="step-box">', unsafe_allow_html=True)
        
        col1, col2 = st.columns(2)
        with col1:
            server = st.text_input("Server", "MAHESHWAGH\\WINCC", key="conn_server")
        
        with col2:
            database = st.text_input("Database", "VPI1", key="conn_database")
        
        # Connect and go to next step button (from XYZ.py logic)
        if st.button("🔗 Connect & Go to Next Step", type="primary", use_container_width=True):
            if not server or not database:
                st.error("Please enter server and database name")
                return
            
            with st.spinner("Connecting..."):
                success, message = st.session_state.db.connect(
                    server=server,
                    database=database
                )
                
                if success:
                    st.success("✅ Connected successfully!")
                    st.session_state.tables_list = st.session_state.db.get_tables()
                    st.session_state.step = 2
                    st.rerun()
                else:
                    st.error(f"❌ {message}")
        
        # Test connection button
        if st.button("🧪 Test Connection", use_container_width=True):
            if not server or not database:
                st.error("Please enter server and database name")
                return
            
            with st.spinner("Testing connection..."):
                success, message = st.session_state.db.connect(
                    server=server,
                    database=database
                )
                
                if success:
                    st.success("✅ Connection test successful!")
                    st.session_state.db.disconnect()
                else:
                    st.error(f"❌ Connection test failed: {message}")
        
        st.markdown('</div>', unsafe_allow_html=True)

def show_table_selection_tab():
    """Table selection tab"""
    st.markdown("## Step 2: Table Selection")
    
    if not st.session_state.db.connected:
        st.warning("Please connect to database first")
        if st.button("← Go to Connection"):
            st.session_state.step = 1
            st.rerun()
        return
    
    with st.container():
        st.markdown('<div class="step-box">', unsafe_allow_html=True)
        
        # Search and filter
        col1, col2, col3 = st.columns([3, 2, 2])
        with col1:
            search = st.text_input("Search tables", placeholder="Type to filter...", key="table_search")
        
        with col2:
            if st.button("Select All", use_container_width=True):
                st.session_state.selected_tables = st.session_state.tables_list.copy()
                st.rerun()
        
        with col3:
            if st.button("Clear All", use_container_width=True):
                st.session_state.selected_tables = []
                st.rerun()
        
        # Filter tables
        filtered = st.session_state.tables_list
        if search:
            filtered = [t for t in filtered if search.lower() in t.lower()]
        
        # Multi-select with checkboxes (like Tkinter)
        selected = st.multiselect(
            "Select tables to export:",
            filtered,
            default=st.session_state.selected_tables,
            help="Select one or more tables",
            key="table_multiselect"
        )
        
        st.session_state.selected_tables = selected
        
        # Show selected tables count
        if selected:
            st.success(f"✅ Selected {len(selected)} table(s)")
            
            with st.expander("📋 View selected tables", expanded=False):
                for i, table in enumerate(selected, 1):
                    display_name = st.session_state.db.get_display_name(table)
                    col1, col2 = st.columns([3, 2])
                    with col1:
                        st.write(f"{i}. **{table}**")
                    with col2:
                        st.write(f"({display_name})")
        else:
            st.info("No tables selected")
        
        st.markdown('</div>', unsafe_allow_html=True)
    
    # Navigation buttons
    col1, col2 = st.columns(2)
    with col1:
        if st.button("← Previous: Connection", use_container_width=True):
            st.session_state.step = 1
            st.rerun()
    with col2:
        if st.button("Next: Position Mapping →", type="primary", use_container_width=True):
            if not st.session_state.selected_tables:
                st.warning("Please select at least one table")
            else:
                st.session_state.step = 3
                st.rerun()

def show_position_mapping_tab():
    """Position mapping tab"""
    st.markdown("## Step 3: Position Mapping")
    
    if not st.session_state.selected_tables:
        st.warning("Please select tables first")
        if st.button("← Go to Table Selection"):
            st.session_state.step = 2
            st.rerun()
        return
    
    with st.container():
        st.markdown('<div class="step-box">', unsafe_allow_html=True)
        
        # Template upload section
        st.markdown("### 📄 Template Selection")
        col1, col2 = st.columns([3, 1])
        with col1:
            uploaded = st.file_uploader("Upload Excel template", type=['xlsx', 'xls'], key="template_upload")
        
        with col2:
            if st.button("Clear Template", use_container_width=True):
                st.session_state.template_path = None
                st.session_state.template_sheets = []
                st.rerun()
        
        if uploaded:
            # Save template
            temp_dir = tempfile.gettempdir()
            template_path = os.path.join(temp_dir, uploaded.name)
            
            with open(template_path, "wb") as f:
                f.write(uploaded.getbuffer())
            
            st.session_state.template_path = template_path
            
            # Get sheet names
            try:
                wb = load_workbook(template_path, read_only=True)
                st.session_state.template_sheets = wb.sheetnames
                st.success(f"✅ Template loaded with {len(st.session_state.template_sheets)} sheet(s)")
                
                with st.expander("View sheets"):
                    for sheet in st.session_state.template_sheets:
                        st.write(f"• {sheet}")
            except Exception as e:
                st.error(f"Error reading template: {e}")
                st.session_state.template_sheets = []
        elif st.session_state.template_path:
            st.info(f"Template: {os.path.basename(st.session_state.template_path)}")
            if st.session_state.template_sheets:
                st.write(f"Sheets: {', '.join(st.session_state.template_sheets[:5])}")
                if len(st.session_state.template_sheets) > 5:
                    st.write(f"... and {len(st.session_state.template_sheets) - 5} more sheets")
        
        # Merge rules section (from XYZ.py)
        st.markdown("### 🔗 Merge Cell Rules (Optional)")
        st.markdown('<div class="merge-rules">', unsafe_allow_html=True)
        
        merge_rules_text = st.text_area(
            "Enter merge ranges (one per line): SheetName!StartCell:EndCell",
            height=100,
            placeholder="Example: Sheet1!B4:D4  (merges B4, C4, D4)\nExample: Sheet1!A1:C1  (merges A1, B1, C1)",
            key="merge_rules_text"
        )
        
        if merge_rules_text:
            st.session_state.merge_rules = [line.strip() for line in merge_rules_text.splitlines() if line.strip()]
            if st.session_state.merge_rules:
                st.info(f"Added {len(st.session_state.merge_rules)} merge rules")
        
        st.markdown('</div>', unsafe_allow_html=True)
        
        # Instructions
        st.markdown("### 📝 Instructions")
        st.markdown("""
        **TWO TYPES OF CONFIGURATION:**
        
        1. **For BACKGROUND_DATA and BATCH_DATA tables:**
           - You only need to specify a starting cell position
           - All data will be inserted starting from this cell
           - Template already has headers - only values inserted
        
        2. **For HEADER tables (and similar):**
           - You need to map each column to a specific cell
           - Only the first row of data will be used
           - Useful for header/static information
        """)
        
        # Configure positions button
        if st.button("⚙️ Configure Position Mappings", type="primary", use_container_width=True):
            if not st.session_state.template_path:
                st.error("Please upload a template first")
                return
            
            # Store merge rules
            if merge_rules_text:
                st.session_state.merge_rules = [line.strip() for line in merge_rules_text.splitlines() if line.strip()]
            
            # Open configuration in new section
            st.session_state.configuring_positions = True
            st.rerun()
        
        # Show current configurations
        if st.session_state.table_configs:
            st.markdown("### 📋 Current Mappings")
            for table_name, config in st.session_state.table_configs.items():
                with st.expander(f"Table: {config.display_name}", expanded=False):
                    if config.start_row > 0 and config.start_col:
                        st.write(f"**Type:** Data Table (BACKGROUND/BATCH)")
                        st.write(f"**Start Position:** {config.start_col}{config.start_row}")
                        if config.apply_to_all_sheets:
                            st.write(f"**Apply to:** All Sheets")
                        elif config.selected_sheets:
                            sheets = config.selected_sheets
                            if len(sheets) <= 3:
                                st.write(f"**Apply to:** {', '.join(sheets)}")
                            else:
                                st.write(f"**Apply to:** {', '.join(sheets[:3])} +{len(sheets)-3} more")
                        else:
                            st.write(f"**Apply to:** {config.sheet_name}")
                    else:
                        st.write(f"**Type:** Header/Static Data")
                        st.write(f"**Total Mappings:** {len(config.column_mappings)}")
                        if config.column_mappings:
                            for col_name, cell_mapping in config.column_mappings.items():
                                if cell_mapping.apply_to_all_sheets:
                                    apply_info = " (All Sheets)"
                                elif cell_mapping.selected_sheets:
                                    sheets = cell_mapping.selected_sheets
                                    if len(sheets) <= 3:
                                        apply_info = f" (Sheets: {', '.join(sheets)})"
                                    else:
                                        apply_info = f" (Sheets: {', '.join(sheets[:3])} +{len(sheets)-3} more)"
                                else:
                                    apply_info = f" (Sheet: {cell_mapping.template_sheet})"
                                st.write(f"• {col_name} → {cell_mapping.template_cell}{apply_info}")
        
        st.markdown('</div>', unsafe_allow_html=True)
    
    # If configuring positions, show configuration interface
    if st.session_state.get('configuring_positions', False):
        show_position_configuration()
    
    # Navigation buttons
    col1, col2 = st.columns(2)
    with col1:
        if st.button("← Previous: Table Selection", use_container_width=True):
            st.session_state.step = 2
            st.rerun()
    with col2:
        if st.button("Next: Filters →", type="primary", use_container_width=True):
            st.session_state.step = 4
            st.rerun()

def show_position_configuration():
    """Show position configuration interface"""
    st.markdown("---")
    st.markdown("### ⚙️ Position Mapping Configuration")
    
    for table_name in st.session_state.selected_tables:
        with st.expander(f"Configure: {table_name}", expanded=False):
            display_name = st.session_state.db.get_display_name(table_name)
            st.write(f"**Display Name:** {display_name}")
            
            # Check table type
            is_simple_table = any(keyword in table_name.upper() for keyword in 
                                 ['BACKGROUND', 'BATCH', 'DATA'])
            
            if is_simple_table:
                # Simple position configuration for BACKGROUND/BATCH tables
                st.write("**Table Type:** Data Table (BACKGROUND/BATCH)")
                
                col1, col2 = st.columns(2)
                with col1:
                    sheet = st.selectbox(
                        f"Target Sheet for {table_name}",
                        st.session_state.template_sheets,
                        key=f"sheet_{table_name}"
                    )
                
                with col2:
                    start_cell = st.text_input(
                        f"Start cell for {table_name}",
                        value="A2",
                        key=f"cell_{table_name}",
                        help="Where should the data start? (e.g., B4, C10)"
                    )
                
                # Apply to options
                apply_option = st.radio(
                    f"Apply to which sheets for {table_name}?",
                    ["This Sheet Only", "All Sheets", "Select Specific Sheets"],
                    key=f"apply_{table_name}"
                )
                
                selected_sheets = []
                if apply_option == "Select Specific Sheets":
                    selected_sheets = st.multiselect(
                        f"Select sheets for {table_name}",
                        st.session_state.template_sheets,
                        default=[sheet],
                        key=f"select_sheets_{table_name}"
                    )
                
                if st.button(f"Save mapping for {table_name}", key=f"save_{table_name}"):
                    if not re.match(r'^[A-Z]+\d+$', start_cell.upper()):
                        st.error("Invalid cell format. Use like B4, C10")
                    else:
                        # Parse cell to get row and column
                        col_letter = ''.join([c for c in start_cell.upper() if c.isalpha()])
                        row_num = int(''.join([c for c in start_cell if c.isdigit()]))
                        
                        # Determine sheets to apply to
                        apply_to_all = (apply_option == "All Sheets")
                        target_sheets = []
                        
                        if apply_option == "This Sheet Only":
                            target_sheets = [sheet]
                        elif apply_option == "All Sheets":
                            target_sheets = st.session_state.template_sheets
                        else:  # Select Specific Sheets
                            target_sheets = selected_sheets
                        
                        st.session_state.table_configs[table_name] = TableConfig(
                            table_name=table_name,
                            display_name=display_name,
                            start_row=row_num,
                            start_col=col_letter,
                            sheet_name=sheet,
                            column_mappings={},
                            apply_to_all_sheets=apply_to_all,
                            selected_sheets=target_sheets
                        )
                        
                        st.success(f"✅ Mapping saved for {table_name}")
            
            else:
                # Column mapping configuration for header tables
                st.write("**Table Type:** Header/Static Data")
                
                # Get columns for mapping
                try:
                    columns = st.session_state.db.get_table_columns(table_name)
                    st.write(f"**Columns found:** {len(columns)}")
                    
                    # Show column mappings in a more structured way
                    for i, col in enumerate(columns[:10]):  # Show first 10 columns
                        col1, col2, col3 = st.columns([3, 2, 1])
                        with col1:
                            st.write(f"**{col}**")
                        with col2:
                            sheet = st.selectbox(
                                f"Sheet for {col}",
                                st.session_state.template_sheets,
                                key=f"sheet_{table_name}_{col}",
                                index=0
                            )
                        with col3:
                            cell = st.text_input(
                                f"Cell for {col}",
                                value="",
                                key=f"cell_{table_name}_{col}",
                                placeholder="e.g., B4"
                            )
                    
                    if len(columns) > 10:
                        st.info(f"... and {len(columns) - 10} more columns")
                    
                    # Apply to options for all columns
                    st.markdown("**Apply to which sheets for all columns?**")
                    apply_option_all = st.radio(
                        f"Apply option for {table_name} columns",
                        ["This Sheet Only", "All Sheets", "Select Specific Sheets"],
                        key=f"apply_all_{table_name}"
                    )
                    
                    selected_sheets_all = []
                    if apply_option_all == "Select Specific Sheets":
                        selected_sheets_all = st.multiselect(
                            f"Select sheets for all columns in {table_name}",
                            st.session_state.template_sheets,
                            default=st.session_state.template_sheets[0],
                            key=f"select_sheets_all_{table_name}"
                        )
                    
                    if st.button(f"Save mappings for {table_name}", key=f"save_mappings_{table_name}"):
                        # Collect all column mappings
                        column_mappings = {}
                        
                        for i, col in enumerate(columns):
                            # Get values for this column
                            sheet_key = f"sheet_{table_name}_{col}"
                            cell_key = f"cell_{table_name}_{col}"
                            
                            # Use default values if not set for all columns
                            sheet_val = st.session_state.get(sheet_key, st.session_state.template_sheets[0] if st.session_state.template_sheets else "")
                            cell_val = st.session_state.get(cell_key, "")
                            
                            if cell_val and re.match(r'^[A-Z]+\d+$', cell_val.upper()):
                                # Determine sheets to apply to
                                apply_all = (apply_option_all == "All Sheets")
                                target_sheets = []
                                
                                if apply_option_all == "This Sheet Only":
                                    target_sheets = [sheet_val]
                                elif apply_option_all == "All Sheets":
                                    target_sheets = st.session_state.template_sheets
                                else:  # Select Specific Sheets
                                    target_sheets = selected_sheets_all
                                
                                column_mappings[col] = CellMapping(
                                    table_name=table_name,
                                    column_name=col,
                                    template_sheet=sheet_val,
                                    template_cell=cell_val.upper(),
                                    apply_to_all_sheets=apply_all,
                                    selected_sheets=target_sheets
                                )
                        
                        if column_mappings:
                            st.session_state.table_configs[table_name] = TableConfig(
                                table_name=table_name,
                                display_name=display_name,
                                start_row=0,  # No start position for header tables
                                start_col="",  # No start column for header tables
                                sheet_name=st.session_state.template_sheets[0] if st.session_state.template_sheets else "",
                                column_mappings=column_mappings,
                                apply_to_all_sheets=(apply_option_all == "All Sheets"),
                                selected_sheets=selected_sheets_all if apply_option_all == "Select Specific Sheets" else []
                            )
                            
                            st.success(f"✅ {len(column_mappings)} column mappings saved for {table_name}")
                        else:
                            st.warning("No valid column mappings found. Please enter at least one valid cell reference.")
                        
                except Exception as e:
                    st.error(f"Error getting columns: {e}")

def show_filters_tab():
    """Filters tab"""
    st.markdown("## Step 4: Data Filters")
    
    if not st.session_state.selected_tables:
        st.warning("Please select tables first")
        if st.button("← Go to Table Selection"):
            st.session_state.step = 2
            st.rerun()
        return
    
    with st.container():
        st.markdown('<div class="step-box">', unsafe_allow_html=True)
        
        st.info("Configure filters to limit exported data for each table")
        
        # Row limit for all tables
        st.markdown("### 📊 Global Settings")
        row_limit = st.number_input(
            "Row limit (0 = all rows)",
            min_value=0,
            max_value=10000,
            value=1000,
            key="global_row_limit"
        )
        st.session_state.row_limit = row_limit
        
        # Configure filters for each table
        for table_name in st.session_state.selected_tables:
            with st.expander(f"Filters for {table_name}", expanded=False):
                # Get existing filters
                current_filters = st.session_state.filters.get(table_name, {})
                
                # Get batches
                batches = st.session_state.db.get_batches_from_table(table_name)
                
                if batches:
                    selected_batch = st.selectbox(
                        f"Select Batch for {table_name}",
                        ["All Batches"] + batches,
                        index=0 if not current_filters.get('batch') else 
                               (["All Batches"] + batches).index(current_filters.get('batch')),
                        key=f"batch_{table_name}"
                    )
                else:
                    selected_batch = "All Batches"
                    st.info("No batch column found in this table")
                
                # Time filters
                st.markdown("**Time Range (Optional):**")
                col1, col2 = st.columns(2)
                with col1:
                    enable_time = st.checkbox(f"Enable time filter for {table_name}", 
                                             value=bool(current_filters.get('start_time') or current_filters.get('end_time')),
                                             key=f"enable_time_{table_name}")
                
                if enable_time:
                    col1, col2 = st.columns(2)
                    with col1:
                        # Default to 24 hours ago
                        default_start = datetime.now() - timedelta(hours=24)
                        start_date = st.date_input(
                            f"Start Date for {table_name}",
                            value=current_filters.get('start_time', default_start).date() if current_filters.get('start_time') else default_start.date(),
                            key=f"start_date_{table_name}"
                        )
                        start_time = st.time_input(
                            f"Start Time for {table_name}",
                            value=current_filters.get('start_time', default_start).time() if current_filters.get('start_time') else default_start.time(),
                            key=f"start_time_{table_name}"
                        )
                    
                    with col2:
                        # Default to now
                        default_end = datetime.now()
                        end_date = st.date_input(
                            f"End Date for {table_name}",
                            value=current_filters.get('end_time', default_end).date() if current_filters.get('end_time') else default_end.date(),
                            key=f"end_date_{table_name}"
                        )
                        end_time = st.time_input(
                            f"End Time for {table_name}",
                            value=current_filters.get('end_time', default_end).time() if current_filters.get('end_time') else default_end.time(),
                            key=f"end_time_{table_name}"
                        )
                
                # Save filters button
                if st.button(f"Save filters for {table_name}", key=f"save_filters_{table_name}"):
                    # Combine date and time
                    start_datetime = None
                    end_datetime = None
                    
                    if enable_time:
                        start_datetime = datetime.combine(start_date, start_time)
                        end_datetime = datetime.combine(end_date, end_time)
                    
                    # Save to session state
                    st.session_state.filters[table_name] = {
                        'batch': None if selected_batch == "All Batches" else selected_batch,
                        'start_time': start_datetime,
                        'end_time': end_datetime
                    }
                    
                    st.success(f"✅ Filters saved for {table_name}")
        
        # Show current filters summary
        if st.session_state.filters:
            st.markdown("### ⚙️ Current Filters Summary")
            for table_name, filters in st.session_state.filters.items():
                filter_text = f"**{table_name}**: "
                if filters.get('batch'):
                    filter_text += f"Batch: {filters['batch']} "
                if filters.get('start_time') and filters.get('end_time'):
                    filter_text += f"| Time: {filters['start_time'].strftime('%Y-%m-%d %H:%M')} to {filters['end_time'].strftime('%Y-%m-%d %H:%M')}"
                st.markdown(f'<div class="filter-info">{filter_text}</div>', unsafe_allow_html=True)
        
        st.markdown('</div>', unsafe_allow_html=True)
    
    # Navigation buttons
    col1, col2 = st.columns(2)
    with col1:
        if st.button("← Previous: Position Mapping", use_container_width=True):
            st.session_state.step = 3
            st.rerun()
    with col2:
        if st.button("Next: Export →", type="primary", use_container_width=True):
            st.session_state.step = 5
            st.rerun()

def show_export_tab():
    """Export tab"""
    st.markdown("## Step 5: Export")
    
    if not st.session_state.selected_tables:
        st.warning("Please select tables first")
        if st.button("← Go to Table Selection"):
            st.session_state.step = 2
            st.rerun()
        return
    
    with st.container():
        st.markdown('<div class="step-box">', unsafe_allow_html=True)
        
        # Configuration summary
        st.info("### 📊 Export Configuration Summary")
        
        summary = f"""
        **Selected Tables:** {len(st.session_state.selected_tables)}
        **Template:** {os.path.basename(st.session_state.template_path) if st.session_state.template_path else 'None (New Excel)'}
        **Mappings:** {len(st.session_state.table_configs)}
        **Filters:** {len(st.session_state.filters)}
        **Row Limit:** {st.session_state.row_limit if st.session_state.row_limit > 0 else 'All'}
        **Merge Rules:** {len(st.session_state.merge_rules)}
        """
        st.markdown(summary)
        
        # Export options
        st.markdown("### 📤 Export Options")
        
        if st.session_state.template_path:
            export_mode = st.radio(
                "Select export mode:",
                ["Use Template (with mappings)", "New Excel File"],
                key="export_mode"
            )
        else:
            export_mode = "New Excel File"
            st.info("No template loaded - will create new Excel file")
        
        # Output filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        default_name = f"export_{timestamp}.xlsx"
        
        output_filename = st.text_input(
            "Output filename",
            value=default_name,
            help="Name for the exported file"
        )
        
        # Preview button
        if st.button("👁️ Preview Data (First 10 Rows)", use_container_width=True):
            with st.spinner("Fetching preview data..."):
                preview_data = {}
                
                for table_name in st.session_state.selected_tables[:3]:  # Limit to 3 tables
                    filters = st.session_state.filters.get(table_name, {})
                    
                    result = st.session_state.db.fetch_filtered_data(
                        table_name=table_name,
                        batch_name=filters.get('batch'),
                        start_time=filters.get('start_time'),
                        end_time=filters.get('end_time'),
                        limit=10  # Preview limit
                    )
                    
                    if result.get('success'):
                        preview_data[table_name] = result
                
                if preview_data:
                    for table_name, data in preview_data.items():
                        with st.expander(f"Preview: {table_name} ({data.get('row_count', 0)} rows)", expanded=False):
                            if data.get('data'):
                                # Create DataFrame with generic column names (since we don't have actual column names)
                                df_preview = pd.DataFrame(
                                    data['data'],
                                    columns=[f"Column_{i+1}" for i in range(len(data['data'][0]) if data['data'] else 0)]
                                )
                                st.dataframe(df_preview, use_container_width=True)
                                
                                # Show filter info
                                if data.get('filters_applied'):
                                    filters = data['filters_applied']
                                    filter_info = []
                                    if filters.get('batch'):
                                        filter_info.append(f"Batch: {filters['batch']}")
                                    if filters.get('start_time'):
                                        filter_info.append(f"Start: {filters['start_time'].strftime('%Y-%m-%d %H:%M')}")
                                    if filters.get('end_time'):
                                        filter_info.append(f"End: {filters['end_time'].strftime('%Y-%m-%d %H:%M')}")
                                    if filter_info:
                                        st.info("Filters applied: " + " | ".join(filter_info))
                else:
                    st.warning("No data to preview")
        
        # Export button - FIXED VERSION
        if st.button("🚀 Export to Excel", type="primary", use_container_width=True):
            if not output_filename.endswith('.xlsx'):
                output_filename += '.xlsx'
            
            temp_dir = tempfile.gettempdir()
            output_path = os.path.join(temp_dir, output_filename)
            
            # Create a placeholder for export log
            export_log_placeholder = st.empty()
            
            def update_export_log(message):
                export_log_placeholder.markdown(f'<div class="export-log">{message}</div>', unsafe_allow_html=True)
            
            # Start export process
            with st.spinner("Starting export process..."):
                try:
                    update_export_log("="*60 + "\n[START] STARTING EXPORT PROCESS\n" + "="*60)
                    
                    # Check if we need template but don't have mappings
                    if export_mode == "Use Template (with mappings)":
                        if not st.session_state.template_path:
                            st.error("No template loaded. Please upload a template first.")
                            return
                        elif not st.session_state.table_configs:
                            st.warning("No position mappings configured.")
                            if not st.checkbox("Continue export without mappings?"):
                                return
                    
                    # Get row limit
                    row_limit = st.session_state.row_limit if st.session_state.row_limit > 0 else None
                    update_export_log(f"Row limit: {row_limit if row_limit else 'All'}")
                    
                    # Fetch data with filters
                    update_export_log("[FETCH] FETCHING FILTERED DATA FROM DATABASE...")
                    all_data = {}
                    
                    progress_bar = st.progress(0)
                    for i, table_name in enumerate(st.session_state.selected_tables):
                        update_export_log(f"Fetching data for {table_name}...")
                        
                        filters = st.session_state.filters.get(table_name, {})
                        
                        result = st.session_state.db.fetch_filtered_data(
                            table_name=table_name,
                            batch_name=filters.get('batch'),
                            start_time=filters.get('start_time'),
                            end_time=filters.get('end_time'),
                            limit=row_limit
                        )
                        
                        all_data[table_name] = result
                        progress_bar.progress((i + 1) / len(st.session_state.selected_tables))
                    
                    # Check if we got any data
                    tables_with_data = [t for t, d in all_data.items() if d.get('success') and d.get('row_count', 0) > 0]
                    if not tables_with_data:
                        st.error("[WARNING] No data found for any table!")
                        update_export_log("[WARNING] No data found for any table!")
                        return
                    
                    update_export_log(f"[OK] Data fetched for {len(tables_with_data)}/{len(all_data)} tables")
                    update_export_log(f"Total rows to export: {sum(d.get('row_count', 0) for d in all_data.values())}")
                    
                    # Export based on mode
                    success = False
                    
                    if export_mode == "Use Template (with mappings)":
                        update_export_log("[EXPORT] EXPORTING TO TEMPLATE...")
                        update_export_log(f"Template: {os.path.basename(st.session_state.template_path)}")
                        update_export_log(f"Mappings: {len(st.session_state.table_configs)}")
                        update_export_log(f"Merge rules: {len(st.session_state.merge_rules)}")
                        
                        success = ExcelTableExporter.export_tables_to_template(
                            tables_data=all_data,
                            template_path=st.session_state.template_path,
                            table_configs=st.session_state.table_configs,
                            output_path=output_path,
                            merge_rules=st.session_state.merge_rules
                        )
                    else:  # New Excel File
                        update_export_log("[EXPORT] CREATING NEW EXCEL FILE...")
                        success = ExcelTableExporter.export_tables_to_new_excel(
                            tables_data=all_data,
                            output_path=output_path
                        )
                    
                    if success:
                        update_export_log("="*60 + "\n[OK] EXPORT COMPLETED SUCCESSFULLY!\n" + "="*60)
                        update_export_log(f"[FILE] File saved as: {output_path}")
                        
                        st.success(f"✅ Export successful: {output_filename}")
                        
                        # Provide download link
                        with open(output_path, 'rb') as f:
                            excel_data = f.read()
                        
                        b64 = base64.b64encode(excel_data).decode()
                        download_link = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{output_filename}" class="stDownloadButton">📥 Download Excel File</a>'
                        st.markdown(download_link, unsafe_allow_html=True)
                        
                        # Display file info
                        file_size = os.path.getsize(output_path) / 1024  # KB
                        st.info(f"**File Size:** {file_size:.2f} KB")
                        st.info(f"**Total Rows Exported:** {sum(data.get('row_count', 0) for data in all_data.values())}")
                        
                        # Debug info
                        with st.expander("📊 Export Details"):
                            export_details = {
                                "tables_exported": len(tables_with_data),
                                "rows_exported": sum(d.get('row_count', 0) for d in all_data.values()),
                                "export_mode": export_mode,
                                "output_path": output_path,
                                "filters_applied": len(st.session_state.filters),
                                "mappings_used": len(st.session_state.table_configs)
                            }
                            st.json(export_details)
                        
                        # Cleanup temp file after download
                        try:
                            os.remove(output_path)
                        except:
                            pass
                    else:
                        st.error("❌ Export failed. Check logs for details.")
                        update_export_log("[ERROR] Export failed. Check 'excel_exporter.log' for details.")
                        
                except Exception as e:
                    error_msg = str(e)
                    st.error(f"❌ Export failed: {error_msg}")
                    update_export_log(f"[ERROR] Export failed: {error_msg}")
                    update_export_log(traceback.format_exc())
                    
                    with st.expander("Error Details"):
                        st.error(f"Failed to export:\n{error_msg}")
                        st.info("""
                        Check the following:
                        1. Close the template file in Excel if it's open
                        2. Verify database connection
                        3. Check position mappings are valid
                        4. Ensure template file is not read-only
                        5. Check 'excel_exporter.log' for detailed error information
                        """)
        
        # Log area
        st.markdown("### 📝 Export Log")
        st.info("Log output will appear above during export process. For detailed logs, check the 'excel_exporter.log' file.")
        
        st.markdown('</div>', unsafe_allow_html=True)
    
    # Navigation buttons
    col1, col2, col3 = st.columns(3)
    with col1:
        if st.button("← Previous: Filters", use_container_width=True):
            st.session_state.step = 4
            st.rerun()
    with col2:
        if st.button("🔄 Start New Export", use_container_width=True):
            # Reset export-related state
            st.session_state.filters = {}
            st.session_state.table_configs = {}
            st.session_state.template_path = None
            st.session_state.template_sheets = []
            st.session_state.merge_rules = []
            st.session_state.step = 2
            st.rerun()
    with col3:
        if st.button("🏠 Go to Connection", use_container_width=True):
            st.session_state.step = 1
            st.rerun()

if __name__ == "__main__":
    main()